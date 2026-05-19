import json
import os
import re
import secrets
import unicodedata
import uuid
from io import BytesIO
from datetime import datetime, date, timedelta
from pathlib import Path
from urllib import request as urlrequest
from urllib.error import HTTPError, URLError

from fastapi import FastAPI, Request, Form, File, UploadFile
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import or_, case, inspect, text, func
from starlette.middleware.sessions import SessionMiddleware
from passlib.context import CryptContext
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse

from .db import SessionLocal, engine, Base
from .models import AppSetting, Order, User, ShowcaseItem, OperationLog, PrintJob, PrintTemplateRule

SECRET_KEY = os.getenv("SECRET_KEY", "dev-secret-key")

app = FastAPI()
app.mount("/static", StaticFiles(directory="app/static"), name="static")
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)
templates = Jinja2Templates(directory="app/templates")

Base.metadata.create_all(bind=engine)


def ensure_schema():
    inspector = inspect(engine)
    if "showcase_items" in inspector.get_table_names():
        columns = {column["name"] for column in inspector.get_columns("showcase_items")}
        if "category" not in columns:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE showcase_items ADD COLUMN category VARCHAR DEFAULT '未分类'"))
        if "item_code" not in columns:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE showcase_items ADD COLUMN item_code VARCHAR DEFAULT ''"))
        with engine.begin() as conn:
            rows = conn.execute(text(
                "SELECT id, COALESCE(category, '未分类') AS category "
                "FROM showcase_items "
                "WHERE item_code IS NULL OR item_code = '' "
                "ORDER BY COALESCE(category, '未分类') ASC, created_at ASC, id ASC"
            )).fetchall()
            counter_rows = conn.execute(text(
                "SELECT COALESCE(category, '未分类') AS category, COUNT(*) AS item_count "
                "FROM showcase_items "
                "WHERE item_code IS NOT NULL AND item_code != '' "
                "GROUP BY COALESCE(category, '未分类')"
            )).fetchall()
            counters = {
                (row._mapping["category"] or "未分类").strip() or "未分类": row._mapping["item_count"]
                for row in counter_rows
            }
            for row in rows:
                category = (row._mapping["category"] or "未分类").strip() or "未分类"
                counters[category] = counters.get(category, 0) + 1
                conn.execute(
                    text("UPDATE showcase_items SET item_code = :code WHERE id = :id"),
                    {"code": f"{category}-{counters[category]:03d}", "id": row._mapping["id"]}
                )
    if "orders" in inspector.get_table_names():
        with engine.begin() as conn:
            legacy_payment_count = conn.execute(text(
                "SELECT COUNT(*) FROM orders WHERE payment_status = '部分付款'"
            )).scalar() or 0
            if legacy_payment_count:
                conn.execute(text("UPDATE orders SET payment_status = '未付款' WHERE payment_status = '部分付款'"))

            stale_unpaid_count = conn.execute(text(
                "SELECT COUNT(*) FROM orders "
                "WHERE unpaid_amount != "
                "CASE "
                "WHEN payment_status = '已付款' THEN 0 "
                "WHEN total_amount - paid_amount > 0 THEN total_amount - paid_amount "
                "ELSE 0 END"
            )).scalar() or 0
            if stale_unpaid_count:
                conn.execute(text(
                    "UPDATE orders SET unpaid_amount = "
                    "CASE "
                    "WHEN payment_status = '已付款' THEN 0 "
                    "WHEN total_amount - paid_amount > 0 THEN total_amount - paid_amount "
                    "ELSE 0 END"
                ))
    if "print_jobs" in inspector.get_table_names():
        columns = {column["name"] for column in inspector.get_columns("print_jobs")}
        if "print_template" not in columns:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE print_jobs ADD COLUMN print_template VARCHAR DEFAULT 'delivery'"))


ensure_schema()

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


# ========================
# 基础工具函数
# ========================

def is_logged_in(request: Request) -> bool:
    return bool(request.session.get("user"))


def get_current_username(request: Request):
    return request.session.get("user")


def require_login(request: Request):
    if not is_logged_in(request):
        return RedirectResponse(url="/login", status_code=303)
    return None


def get_current_user(request: Request, db: Session):
    username = request.session.get("user")
    if not username:
        return None
    return db.query(User).filter(User.username == username).first()


def require_admin(request: Request, db: Session):
    redirect = require_login(request)
    if redirect:
        return redirect

    current_user = get_current_user(request, db)
    if current_user is None or not current_user.is_admin:
        return RedirectResponse(url="/orders", status_code=303)
    return None


def add_flash(request: Request, message: str, level: str = "success"):
    flashes = list(request.session.get("flashes", []))
    flashes.append({"message": message, "level": level})
    request.session["flashes"] = flashes[-5:]


def get_flashes(request: Request):
    if hasattr(request.state, "flashes"):
        return request.state.flashes
    flashes = request.session.pop("flashes", [])
    request.state.flashes = flashes
    return flashes


templates.env.globals["get_flashes"] = get_flashes


def hash_password(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain_password: str, password_hash: str) -> bool:
    return pwd_context.verify(plain_password, password_hash)


def generate_order_no(db: Session) -> str:
    today_prefix = datetime.now().strftime("%Y%m%d")
    like_pattern = f"{today_prefix}-%"

    today_orders = (
        db.query(Order)
        .filter(Order.order_no.like(like_pattern))
        .order_by(Order.order_no.desc())
        .all()
    )

    if not today_orders:
        return f"{today_prefix}-001"

    last_order_no = today_orders[0].order_no
    last_seq = int(last_order_no.split("-")[-1])
    return f"{today_prefix}-{last_seq + 1:03d}"


def calc_payment_status(total_amount: float, deposit_amount: float, payment_status: str = "未付款"):
    deposit_amount = max(deposit_amount or 0, 0)
    balance_due = max((total_amount or 0) - deposit_amount, 0)

    if payment_status == "已付款" or (total_amount or 0) > 0 and deposit_amount >= total_amount:
        return 0.0, "已付款"
    return balance_due, "未付款"


def parse_date_field(value: str, label: str):
    value = str(value or "").strip()
    if not value:
        return None, ""
    try:
        return date.fromisoformat(value), ""
    except ValueError:
        return None, f"{label}格式不正确，请使用日期选择器重新选择"


def order_form_error(
    customer: str,
    item_name: str,
    quantity: int,
    unit_price: float,
    paid_amount: float,
    priority_color: str,
    due_date: str,
):
    errors = []
    customer = str(customer or "").strip()
    item_name = str(item_name or "").strip()
    if not customer:
        errors.append("客户名称不能为空")
    if not item_name:
        errors.append("商品名称不能为空")
    if quantity <= 0:
        errors.append("数量必须大于 0")
    if unit_price <= 0:
        errors.append("单价必须大于 0")
    if paid_amount < 0:
        errors.append("已收金额不能小于 0")

    total_amount = max(quantity, 0) * max(unit_price, 0)
    if total_amount and paid_amount > total_amount:
        errors.append("已收金额不能大于总金额")

    if priority_color not in ["红色", "橙色", "黄色", "蓝色", "灰色"]:
        errors.append("优先级颜色不正确")

    parsed_due_date, date_error = parse_date_field(due_date, "截至日期")
    if date_error:
        errors.append(date_error)

    return errors, parsed_due_date


def payment_snapshot(order: Order):
    return {
        "payment_status": order.payment_status,
        "paid_amount": float(order.paid_amount or 0),
        "unpaid_amount": float(order.unpaid_amount or 0),
    }


def parse_payment_snapshot(value: str):
    try:
        data = json.loads(value)
    except (TypeError, ValueError, json.JSONDecodeError):
        return {"payment_status": str(value or "未付款")}
    if not isinstance(data, dict):
        return {"payment_status": "未付款"}
    return data


def action_label(action: str):
    labels = {
        "queue_print": "加入打印队列",
        "mark_printed": "打印成功",
        "print_failed": "打印失败",
        "retry_print": "重新打印",
        "mark_paid": "标记结清",
        "mark_unpaid": "改回未付款",
        "undo_payment_status": "撤回付款状态",
        "undo_print_status": "撤回打印状态",
        "edit_order": "编辑订单",
        "delete_order": "删除订单",
    }
    return labels.get(action, action)


templates.env.globals["action_label"] = action_label


def display_width(value: str) -> int:
    width = 0
    for char in str(value):
        width += 2 if unicodedata.east_asian_width(char) in ("W", "F") else 1
    return width


def fit_display(value, width: int, align: str = "left") -> str:
    text_value = str(value or "")
    result = ""
    used = 0
    for char in text_value:
        char_width = 2 if unicodedata.east_asian_width(char) in ("W", "F") else 1
        if used + char_width > width:
            break
        result += char
        used += char_width

    padding = " " * max(width - used, 0)
    if align == "right":
        return padding + result
    if align == "center":
        left = len(padding) // 2
        return padding[:left] + result + padding[left:]
    return result + padding


def wrap_display(value, width: int) -> list[str]:
    lines = []
    current = ""
    used = 0
    for char in str(value or ""):
        char_width = 2 if unicodedata.east_asian_width(char) in ("W", "F") else 1
        if current and used + char_width > width:
            lines.append(current)
            current = char
            used = char_width
        else:
            current += char
            used += char_width
    if current:
        lines.append(current)
    return lines or [""]


def format_money(value) -> str:
    return f"{float(value or 0):.2f}"


def build_delivery_row(name="", spec="", unit="", quantity="", unit_price="", amount="", remark="") -> str:
    return (
        "|"
        + fit_display(name, 18)
        + "|"
        + fit_display(spec, 12)
        + "|"
        + fit_display(unit, 6, "center")
        + "|"
        + fit_display(quantity, 6, "right")
        + "|"
        + fit_display(unit_price, 8, "right")
        + "|"
        + fit_display(amount, 10, "right")
        + "|"
        + fit_display(remark, 16)
        + "|"
    )


PRINT_TEMPLATES = {
    "delivery": {
        "label": "标准出货单",
        "description": "当前默认模板，适合普通连续纸或 A5/A4 打印。",
    },
    "a4": {
        "label": "A4 出货单",
        "description": "内容更舒展，适合 A4 纸。",
    },
    "duplicate": {
        "label": "二联单",
        "description": "同一订单打印客户联和存根联。",
    },
    "triplicate": {
        "label": "三联单",
        "description": "同一订单打印客户联、财务联和存根联。",
    },
    "receipt": {
        "label": "小票",
        "description": "窄纸简版，只保留核心金额和签收信息。",
    },
}


templates.env.globals["print_templates"] = PRINT_TEMPLATES


def normalize_print_template(template_key: str) -> str:
    template_key = str(template_key or "").strip()
    return template_key if template_key in PRINT_TEMPLATES else "delivery"


def is_auto_print_template(template_key: str) -> bool:
    return str(template_key or "").strip() in ("", "auto")


def print_template_label(template_key: str) -> str:
    return PRINT_TEMPLATES[normalize_print_template(template_key)]["label"]


templates.env.globals["print_template_label"] = print_template_label


def get_app_setting(db: Session, key: str, default: str = "") -> str:
    setting = db.query(AppSetting).filter(AppSetting.key == key).first()
    if setting is None:
        return default
    return str(setting.value or default)


def set_app_setting(db: Session, key: str, value: str):
    setting = db.query(AppSetting).filter(AppSetting.key == key).first()
    if setting is None:
        setting = AppSetting(key=key, value=value, updated_at=datetime.utcnow())
        db.add(setting)
    else:
        setting.value = value
        setting.updated_at = datetime.utcnow()
    return setting


def get_default_print_template(db: Session) -> str:
    return normalize_print_template(get_app_setting(db, "default_print_template", "delivery"))


def get_customer_print_template(db: Session, customer_name: str):
    customer_name = str(customer_name or "").strip()
    if not customer_name:
        return None
    rule = db.query(PrintTemplateRule).filter(PrintTemplateRule.customer_name == customer_name).first()
    if rule is None:
        return None
    return normalize_print_template(rule.print_template)


def resolve_print_template(db: Session, order: Order, requested_template: str = "auto") -> str:
    if not is_auto_print_template(requested_template):
        return normalize_print_template(requested_template)
    customer_template = get_customer_print_template(db, order.customer)
    if customer_template:
        return customer_template
    return get_default_print_template(db)


def build_print_text(order: Order, template_key: str = "delivery") -> str:
    template_key = normalize_print_template(template_key)
    if template_key == "a4":
        return build_a4_print_text(order)
    if template_key == "duplicate":
        return build_multi_copy_print_text(order, ["客户联", "存根联"])
    if template_key == "triplicate":
        return build_multi_copy_print_text(order, ["客户联", "财务联", "存根联"])
    if template_key == "receipt":
        return build_receipt_print_text(order)
    return build_delivery_print_text(order)


def build_delivery_print_text(order: Order, copy_label: str = "") -> str:
    table_width = 86
    line = "+" + "-" * 18 + "+" + "-" * 12 + "+" + "-" * 6 + "+" + "-" * 6 + "+" + "-" * 8 + "+" + "-" * 10 + "+" + "-" * 16 + "+"
    customer = f"客户名称：{order.customer or ''}"
    item_lines = wrap_display(order.item_name or "", 18)
    spec_lines = wrap_display(order.size or "", 12)
    remark_lines = wrap_display(order.remark or "", 16)
    row_count = max(len(item_lines), len(spec_lines), len(remark_lines), 1)

    lines = [
        fit_display(f"方圆五金出货单{('（' + copy_label + '）') if copy_label else ''}", table_width, "center"),
        fit_display(f"订单号：{order.order_no}", table_width, "right"),
        "",
        customer,
        line,
        build_delivery_row("货物名称", "规格", "单位", "数量", "单价", "货款", "备注"),
        line,
    ]

    for index in range(row_count):
        lines.append(build_delivery_row(
            item_lines[index] if index < len(item_lines) else "",
            spec_lines[index] if index < len(spec_lines) else "",
            "件" if index == 0 else "",
            order.quantity if index == 0 else "",
            format_money(order.unit_price) if index == 0 else "",
            format_money(order.total_amount) if index == 0 else "",
            remark_lines[index] if index < len(remark_lines) else "",
        ))

    for _ in range(max(5 - row_count, 0)):
        lines.append(build_delivery_row())

    lines.extend([
        line,
        build_delivery_row("合计", "", "", order.quantity, "", format_money(order.total_amount), ""),
        line,
        "",
    ])

    footer_text = (
        "本厂大型激光切割，剪板，折叠对外加工，专业定尺生产瓦楞板，波浪板，"
        "三角板，不锈钢板，黑钛瓦楞，三角，波浪板均可"
    )
    lines.extend(wrap_display(footer_text, table_width))
    lines.extend([
        "定尺生产！可提供火锅桌架全套配件！",
        "门市地址：东段君良仓储西800米路北  厂址：东段速8酒店后面50米道东",
        "电话：15226662348    13582962755（微信同步）",
        "送货人：厂内司机     收货人签字：",
        "",
        "\n",
    ])
    return "\n".join(lines)


def build_multi_copy_print_text(order: Order, copy_labels: list[str]) -> str:
    return ("\n" + "=" * 86 + "\n").join(
        build_delivery_print_text(order, label).rstrip()
        for label in copy_labels
    ) + "\n"


def build_a4_print_text(order: Order) -> str:
    width = 96
    separator = "=" * width
    lines = [
        fit_display("方圆五金出货单", width, "center"),
        fit_display(f"订单号：{order.order_no}    日期：{order.created_at.strftime('%Y-%m-%d') if order.created_at else ''}", width, "right"),
        separator,
        f"客户：{order.customer or ''}",
        f"电话：{order.phone or ''}",
        f"商品：{order.item_name or ''}",
        f"规格：{order.size or ''}",
        f"数量：{order.quantity or 0}    单价：{format_money(order.unit_price)}    货款：{format_money(order.total_amount)}",
        f"已收：{format_money(order.paid_amount)}    待结：{format_money(order.unpaid_amount)}    付款状态：{order.payment_status or ''}",
        separator,
        "备注：",
    ]
    lines.extend(wrap_display(order.remark or "无", width))
    lines.extend([
        separator,
        "本厂大型激光切割，剪板，折叠对外加工，专业定尺生产瓦楞板，波浪板，三角板等。",
        "门市地址：东段君良仓储西800米路北  厂址：东段速8酒店后面50米道东",
        "电话：15226662348    13582962755（微信同步）",
        "",
        "送货人：厂内司机                              收货人签字：",
        "",
        "\n",
    ])
    return "\n".join(lines)


def build_receipt_print_text(order: Order) -> str:
    width = 42
    line = "-" * width
    lines = [
        fit_display("方圆五金出货单", width, "center"),
        line,
        f"单号：{order.order_no}",
        f"客户：{order.customer or ''}",
        f"电话：{order.phone or ''}",
        line,
    ]
    lines.extend(wrap_display(f"商品：{order.item_name or ''}", width))
    lines.extend(wrap_display(f"规格：{order.size or ''}", width))
    lines.extend([
        f"数量：{order.quantity or 0}",
        f"单价：{format_money(order.unit_price)}",
        f"金额：{format_money(order.total_amount)}",
        f"已收：{format_money(order.paid_amount)}",
        f"待结：{format_money(order.unpaid_amount)}",
        line,
    ])
    if order.remark:
        lines.extend(wrap_display(f"备注：{order.remark}", width))
        lines.append(line)
    lines.extend([
        "电话：15226662348",
        "      13582962755",
        "收货人签字：",
        "",
        "\n",
    ])
    return "\n".join(lines)


def verify_print_client(request: Request) -> bool:
    expected_token = os.getenv("PRINT_CLIENT_TOKEN", "").strip()
    supplied_token = (
        request.headers.get("X-Print-Client-Token")
        or request.query_params.get("token")
        or ""
    ).strip()
    return bool(expected_token) and secrets.compare_digest(supplied_token, expected_token)


def get_active_print_job(db: Session, order_id: int):
    return (
        db.query(PrintJob)
        .filter(
            PrintJob.order_id == order_id,
            PrintJob.status.in_(["pending", "printing"])
        )
        .order_by(PrintJob.id.desc())
        .first()
    )


def queue_print_jobs(
    db: Session,
    orders: list[Order],
    operator: str = "",
    print_template: str = "auto",
):
    queued = []
    skipped = []

    for order in orders:
        if order.print_status == "已打印":
            skipped.append(order)
            continue
        active_job = get_active_print_job(db, order.id)
        if active_job is not None:
            skipped.append(order)
            continue

        resolved_template = resolve_print_template(db, order, print_template)
        job = PrintJob(order_id=order.id, print_template=resolved_template, status="pending")
        db.add(job)
        db.flush()

        log_operation(
            db=db,
            target_type="order",
            target_id=order.id,
            action="queue_print",
            field_name="print_job",
            old_value="",
            new_value=f"job_id={job.id};template={resolved_template}",
            operator=operator
        )
        queued.append(order)

    return queued, skipped

def log_operation(
    db: Session,
    target_type: str,
    target_id: int,
    action: str,
    field_name: str,
    old_value: str,
    new_value: str,
    operator: str
):
    log = OperationLog(
        target_type=target_type,
        target_id=target_id,
        action=action,
        field_name=field_name,
        old_value=str(old_value),
        new_value=str(new_value),
        operator=operator
    )
    db.add(log)


def get_latest_reversible_log(db: Session, order_id: int):
    reversible_actions = ["mark_printed", "mark_production", "mark_complete", "mark_paid", "mark_unpaid"]
    undo_actions = ["undo_print_status", "undo_production_status", "undo_payment_status"]

    latest_undo = (
        db.query(OperationLog)
        .filter(
            OperationLog.target_type == "order",
            OperationLog.target_id == order_id,
            OperationLog.action.in_(undo_actions)
        )
        .order_by(OperationLog.id.desc())
        .first()
    )

    query = (
        db.query(OperationLog)
        .filter(
            OperationLog.target_type == "order",
            OperationLog.target_id == order_id,
            OperationLog.action.in_(reversible_actions)
        )
    )
    if latest_undo is not None:
        query = query.filter(OperationLog.id > latest_undo.id)

    return query.order_by(OperationLog.id.desc()).first()

def get_priority_rank_expr():
    return case(
        (Order.priority_color == "红色", 0),
        (Order.priority_color == "橙色", 1),
        (Order.priority_color == "黄色", 2),
        (Order.priority_color == "蓝色", 3),
        else_=4
    )


def get_recent_distinct_values(db: Session, column, keyword: str = "", limit: int = 20):
    query = db.query(
        column.label("value"),
        func.max(Order.created_at).label("last_used_at"),
        func.max(Order.id).label("last_order_id")
    ).filter(column != None, column != "")

    if keyword:
        query = query.filter(column.like(f"%{keyword}%"))

    rows = (
        query
        .group_by(column)
        .order_by(func.max(Order.created_at).desc(), func.max(Order.id).desc())
        .limit(limit)
        .all()
    )
    return [row.value for row in rows if row.value]


def get_showcase_category_options(db: Session):
    return [
        row[0]
        for row in db.query(ShowcaseItem.category)
        .filter(ShowcaseItem.category != None, ShowcaseItem.category != "")
        .distinct()
        .order_by(ShowcaseItem.category.asc())
        .all()
        if row[0]
    ]


def generate_showcase_item_code(db: Session, category: str) -> str:
    category_value = category.strip() or "未分类"
    existing_count = (
        db.query(ShowcaseItem)
        .filter(ShowcaseItem.category == category_value)
        .count()
    )
    return f"{category_value}-{existing_count + 1:03d}"


def safe_redirect_path(return_to: str, default: str) -> str:
    return_to = str(return_to or "")
    if return_to.startswith("/") and not return_to.startswith("//"):
        return return_to
    return default


def get_print_status_label(status: str):
    labels = {
        "pending": "等待打印",
        "printing": "打印中",
        "done": "已完成",
        "failed": "打印失败",
    }
    return labels.get(status, status or "未知")


ORDER_DRAFT_FIELDS = {
    "customer",
    "phone",
    "item_name",
    "size",
    "quantity",
    "unit_price",
    "paid_amount",
    "priority_color",
    "due_date",
    "remark",
}


def normalize_order_draft(raw: dict):
    draft = {field: raw.get(field, "") for field in ORDER_DRAFT_FIELDS}
    issues = []

    for field in ["customer", "phone", "item_name", "size", "remark"]:
        draft[field] = str(draft.get(field) or "").strip()

    try:
        draft["quantity"] = int(draft.get("quantity") or 0)
    except (TypeError, ValueError):
        draft["quantity"] = 0
    if draft["quantity"] <= 0:
        issues.append("数量缺失或不合法")

    for money_field in ["unit_price", "paid_amount"]:
        try:
            draft[money_field] = float(draft.get(money_field) or 0)
        except (TypeError, ValueError):
            draft[money_field] = 0.0
    if draft["unit_price"] <= 0:
        issues.append("单价缺失或不合法")
    if draft["paid_amount"] < 0:
        draft["paid_amount"] = 0.0

    if not draft["customer"]:
        issues.append("客户名称缺失")
    if not draft["item_name"]:
        issues.append("商品名称缺失")

    if draft.get("priority_color") not in ["红色", "橙色", "黄色", "蓝色", "灰色"]:
        draft["priority_color"] = "灰色"

    due_date = str(draft.get("due_date") or "").strip()
    if due_date:
        try:
            date.fromisoformat(due_date)
        except ValueError:
            issues.append("截至日期格式需要是 YYYY-MM-DD")
            due_date = ""
    draft["due_date"] = due_date

    return draft, issues


def build_order_hotwords(db: Session):
    customers = [row[0] for row in db.query(Order.customer).distinct().order_by(Order.customer.asc()).limit(60).all() if row[0]]
    items = [row[0] for row in db.query(Order.item_name).distinct().order_by(Order.item_name.asc()).limit(80).all() if row[0]]
    sizes = [row[0] for row in db.query(Order.size).distinct().order_by(Order.size.asc()).limit(80).all() if row[0]]
    return {"customers": customers, "items": items, "sizes": sizes}


def apply_hotwords_to_draft(text_value: str, draft: dict, hotwords: dict | None):
    if not hotwords:
        return draft

    for field, values in [("customer", hotwords.get("customers", [])), ("item_name", hotwords.get("items", [])), ("size", hotwords.get("sizes", []))]:
        if draft.get(field):
            continue
        matches = [value for value in values if value and value in text_value]
        if matches:
            draft[field] = max(matches, key=len)

    return draft


def fill_recent_unit_price(db: Session, draft: dict):
    if draft.get("unit_price", 0) > 0 or not draft.get("item_name"):
        return draft

    query = db.query(Order).filter(Order.item_name == draft["item_name"])
    if draft.get("size"):
        query = query.filter(Order.size == draft["size"])

    recent_order = query.order_by(Order.created_at.desc(), Order.id.desc()).first()
    if recent_order and recent_order.unit_price:
        draft["unit_price"] = recent_order.unit_price
    return draft


def normalize_order_result(raw: dict, text_value: str, db: Session, hotwords: dict | None):
    raw_orders = raw.get("orders") if isinstance(raw.get("orders"), list) else None
    source_issues = raw.get("issues") if isinstance(raw.get("issues"), list) else []
    confidence = raw.get("confidence", 0.7)

    if not raw_orders:
        raw_orders = [raw]

    orders = []
    all_issues = [str(issue) for issue in source_issues if issue]
    for index, raw_order in enumerate(raw_orders, start=1):
        if not isinstance(raw_order, dict):
            all_issues.append(f"第 {index} 个订单结构不合法")
            continue
        draft, issues = normalize_order_draft(raw_order)
        draft = apply_hotwords_to_draft(text_value, draft, hotwords)
        draft = fill_recent_unit_price(db, draft)
        draft, issues = normalize_order_draft(draft)
        orders.append({"draft": draft, "issues": issues})
        all_issues.extend([f"订单 {index}：{issue}" for issue in issues])

    try:
        confidence = float(confidence)
    except (TypeError, ValueError):
        confidence = 0.7

    return {
        "orders": orders,
        "issues": all_issues,
        "confidence": max(0, min(confidence, 1)),
    }


def heuristic_order_draft(text_value: str, hotwords: dict | None = None):
    text_value = text_value.strip()
    draft = {
        "customer": "",
        "phone": "",
        "item_name": "",
        "size": "",
        "quantity": 1,
        "unit_price": 0,
        "paid_amount": 0,
        "priority_color": "灰色",
        "due_date": "",
        "remark": text_value,
    }

    phone_match = re.search(r"1[3-9]\d{9}", text_value)
    if phone_match:
        draft["phone"] = phone_match.group(0)

    quantity_match = re.search(r"(\d+)\s*(?:个|件|套|张|米|份|只|台|本)?", text_value)
    if quantity_match:
        draft["quantity"] = int(quantity_match.group(1))

    price_match = re.search(r"(?:单价|每个|每件|价格)\s*(\d+(?:\.\d+)?)", text_value)
    if price_match:
        draft["unit_price"] = float(price_match.group(1))

    paid_match = re.search(r"(?:定金|已付|付了)\s*(\d+(?:\.\d+)?)", text_value)
    if paid_match:
        draft["paid_amount"] = float(paid_match.group(1))

    for color in ["红色", "橙色", "黄色", "蓝色", "灰色"]:
        if color in text_value:
            draft["priority_color"] = color
            break
    if "加急" in text_value or "急" in text_value:
        draft["priority_color"] = "红色"

    customer_match = re.search(r"(?:客户|给|帮)\s*([\u4e00-\u9fa5A-Za-z0-9_-]{2,12})", text_value)
    if customer_match:
        draft["customer"] = customer_match.group(1)

    item_match = re.search(r"(?:做|要|下单|订)\s*([\u4e00-\u9fa5A-Za-z0-9_ xX*.-]{2,30})", text_value)
    if item_match:
        draft["item_name"] = item_match.group(1).strip(" ，,。")

    size_match = re.search(r"(\d+(?:\.\d+)?\s*[xX*]\s*\d+(?:\.\d+)?(?:\s*[xX*]\s*\d+(?:\.\d+)?)?)", text_value)
    if size_match:
        draft["size"] = size_match.group(1).replace(" ", "")

    draft = apply_hotwords_to_draft(text_value, draft, hotwords)
    return normalize_order_draft(draft)


def extract_json_object(content: str):
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", content, re.S)
        if not match:
            raise
        return json.loads(match.group(0))


def call_deepseek_model(model: str, user_text: str, hotwords: dict | None = None):
    api_key = os.getenv("DEEPSEEK_API_KEY", "").strip()
    if not api_key:
        return None, "未配置 DEEPSEEK_API_KEY，已使用本地规则生成草稿"

    base_url = os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com/chat/completions").strip()
    system_prompt = (
        "你是订单录入助手。请把用户的自然语言订单解析成 JSON 对象，只返回合法 JSON 字符串，不要 Markdown。"
        "顶层字段必须包含 orders, issues, confidence。orders 是订单数组，每个订单字段必须包含："
        "customer, phone, item_name, size, quantity, unit_price, paid_amount, priority_color, due_date, remark。"
        "paid_amount 表示已收金额，例如定金或已付金额；没有收款就填 0。"
        "priority_color 只能是红色/橙色/黄色/蓝色/灰色；"
        "due_date 使用 YYYY-MM-DD 或空字符串；无法确定的字段用空字符串或 0。\n\n"
        f"历史热词 JSON：{json.dumps(hotwords or {}, ensure_ascii=False)}\n"
        "如果用户说法接近历史热词，优先使用历史热词中的标准写法。\n\n"
        "EXAMPLE INPUT:\n"
        "给张三做亚克力牌 30x40 两个，单价 80，定金 50，明天要，加急。另外给李四做门头一套。\n\n"
        "EXAMPLE JSON OUTPUT:\n"
        "{\n"
        '  "orders": [\n'
        "    {\n"
        '      "customer": "张三",\n'
        '      "phone": "",\n'
        '      "item_name": "亚克力牌",\n'
        '      "size": "30x40",\n'
        '      "quantity": 2,\n'
        '      "unit_price": 80,\n'
        '      "paid_amount": 50,\n'
        '      "priority_color": "红色",\n'
        '      "due_date": "",\n'
        '      "remark": "明天要，加急"\n'
        "    },\n"
        "    {\n"
        '      "customer": "李四",\n'
        '      "phone": "",\n'
        '      "item_name": "门头",\n'
        '      "size": "",\n'
        '      "quantity": 1,\n'
        '      "unit_price": 0,\n'
        '      "paid_amount": 0,\n'
        '      "priority_color": "灰色",\n'
        '      "due_date": "",\n'
        '      "remark": ""\n'
        "    }\n"
        "  ],\n"
        '  "issues": ["第 2 单缺少单价"],\n'
        '  "confidence": 0.82\n'
        "}"
    )
    payload = {
        "model": model,
        "temperature": 0.1,
        "max_tokens": 800,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_text},
        ],
    }
    req = urlrequest.Request(
        base_url,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urlrequest.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
        return None, f"DeepSeek 调用失败：{exc}"

    content = data["choices"][0]["message"].get("content") or ""
    if not content.strip():
        return None, "DeepSeek 返回了空 JSON 内容，请重试或补充订单描述"
    return extract_json_object(content), ""


def call_deepseek_query_model(model: str, user_text: str, hotwords: dict | None = None):
    api_key = os.getenv("DEEPSEEK_API_KEY", "").strip()
    if not api_key:
        return None, "未配置 DEEPSEEK_API_KEY，已使用本地规则生成查询条件"

    base_url = os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com/chat/completions").strip()
    system_prompt = (
        "你是订单查询助手。请把用户的自然语言查询解析成 JSON 对象，只返回合法 JSON 字符串，不要 Markdown。"
        "JSON 字段必须包含：keyword, payment_status, production_status, print_status, date_from, date_to, sort_by, issues, confidence。"
        "payment_status 只能是空字符串/未付款/已付款；production_status 只能是空字符串/未投产/已投产/已完成；"
        "print_status 只能是空字符串/未打印/已打印；date_from/date_to 使用 YYYY-MM-DD 或空字符串；"
        "sort_by 只能是 priority_due/payment_first/order_new/order_old/customer/item/amount_desc/amount_asc/unpaid_desc。\n\n"
        f"今天日期：{date.today().isoformat()}\n"
        f"历史热词 JSON：{json.dumps(hotwords or {}, ensure_ascii=False)}\n"
        "如果用户说到了历史客户、商品或尺寸，把它放进 keyword。\n\n"
        "EXAMPLE INPUT:\n"
        "查一下方圆这个月未付款的订单，待结金额高的排前面\n\n"
        "EXAMPLE JSON OUTPUT:\n"
        "{\n"
        '  "keyword": "方圆",\n'
        '  "payment_status": "未付款",\n'
        '  "production_status": "",\n'
        '  "print_status": "",\n'
        '  "date_from": "2026-04-01",\n'
        '  "date_to": "2026-04-30",\n'
        '  "sort_by": "unpaid_desc",\n'
        '  "issues": [],\n'
        '  "confidence": 0.88\n'
        "}"
    )
    payload = {
        "model": model,
        "temperature": 0.1,
        "max_tokens": 600,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_text},
        ],
    }
    req = urlrequest.Request(
        base_url,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urlrequest.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
        return None, f"DeepSeek 调用失败：{exc}"

    content = data["choices"][0]["message"].get("content") or ""
    if not content.strip():
        return None, "DeepSeek 返回了空 JSON 内容，请重试或补充查询描述"
    return extract_json_object(content), ""


def month_range(base_date: date, offset: int = 0):
    month_index = base_date.month - 1 + offset
    year = base_date.year + month_index // 12
    month = month_index % 12 + 1
    start = date(year, month, 1)
    next_month_index = month_index + 1
    next_year = base_date.year + next_month_index // 12
    next_month = next_month_index % 12 + 1
    end = date(next_year, next_month, 1)
    end = date.fromordinal(end.toordinal() - 1)
    return start.isoformat(), end.isoformat()


def normalize_order_query(raw: dict):
    allowed_payment = ["", "未付款", "已付款"]
    allowed_production = ["", "未投产", "已投产", "已完成"]
    allowed_print = ["", "未打印", "已打印"]
    allowed_sort = ["payment_first", "order_new", "order_old", "customer", "item", "amount_desc", "amount_asc", "unpaid_desc", "priority_due"]

    draft = {
        "keyword": str(raw.get("keyword") or "").strip(),
        "payment_status": str(raw.get("payment_status") or "").strip(),
        "production_status": str(raw.get("production_status") or "").strip(),
        "print_status": str(raw.get("print_status") or "").strip(),
        "date_from": str(raw.get("date_from") or "").strip(),
        "date_to": str(raw.get("date_to") or "").strip(),
        "sort_by": str(raw.get("sort_by") or "priority_due").strip(),
    }
    issues = [str(issue) for issue in raw.get("issues", []) if issue] if isinstance(raw.get("issues"), list) else []

    if draft["payment_status"] not in allowed_payment:
        issues.append("付款状态不明确")
        draft["payment_status"] = ""
    if draft["production_status"] not in allowed_production:
        issues.append("投产状态不明确")
        draft["production_status"] = ""
    if draft["print_status"] not in allowed_print:
        issues.append("打印状态不明确")
        draft["print_status"] = ""
    if draft["sort_by"] not in allowed_sort:
        draft["sort_by"] = "priority_due"

    for field in ["date_from", "date_to"]:
        if draft[field]:
            try:
                date.fromisoformat(draft[field])
            except ValueError:
                issues.append(f"{field} 日期格式不正确")
                draft[field] = ""

    try:
        confidence = float(raw.get("confidence", 0.7))
    except (TypeError, ValueError):
        confidence = 0.7

    return draft, issues, max(0, min(confidence, 1))


def heuristic_order_query(text_value: str, hotwords: dict | None = None):
    today = date.today()
    draft = {
        "keyword": "",
        "payment_status": "",
        "production_status": "",
        "print_status": "",
        "date_from": "",
        "date_to": "",
        "sort_by": "priority_due",
    }

    for status in ["未付款", "已付款"]:
        if status in text_value:
            draft["payment_status"] = status
            break
    for status in ["未投产", "已投产", "已完成"]:
        if status in text_value:
            draft["production_status"] = status
            break
    for status in ["未打印", "已打印"]:
        if status in text_value:
            draft["print_status"] = status
            break

    if "今天" in text_value:
        draft["date_from"] = today.isoformat()
        draft["date_to"] = today.isoformat()
    elif "本月" in text_value or "这个月" in text_value:
        draft["date_from"], draft["date_to"] = month_range(today)
    elif "上月" in text_value or "上个月" in text_value:
        draft["date_from"], draft["date_to"] = month_range(today, -1)
    elif "今年" in text_value or "本年" in text_value:
        draft["date_from"] = date(today.year, 1, 1).isoformat()
        draft["date_to"] = date(today.year, 12, 31).isoformat()

    if ("未付" in text_value or "待结" in text_value) and ("高" in text_value or "多" in text_value):
        draft["sort_by"] = "unpaid_desc"
    elif "金额高" in text_value or "总额高" in text_value:
        draft["sort_by"] = "amount_desc"
    elif "金额低" in text_value or "总额低" in text_value:
        draft["sort_by"] = "amount_asc"
    elif "新" in text_value:
        draft["sort_by"] = "order_new"
    elif "旧" in text_value:
        draft["sort_by"] = "order_old"

    if hotwords:
        values = hotwords.get("customers", []) + hotwords.get("items", []) + hotwords.get("sizes", [])
        matches = [value for value in values if value and value in text_value]
        if matches:
            draft["keyword"] = max(matches, key=len)

    if not draft["keyword"]:
        keyword_match = re.search(r"(?:查|查询|搜索|找|看看)\s*([\u4e00-\u9fa5A-Za-z0-9_*xX.-]{2,20})", text_value)
        if keyword_match:
            draft["keyword"] = keyword_match.group(1).strip(" 的订单")

    return normalize_order_query({**draft, "issues": [], "confidence": 0.45})


SHOWCASE_IMAGE_SIZE = (900, 900)
QUOTE_IMAGE_MAX_BYTES = 4 * 1024 * 1024


def get_pillow_tools():
    try:
        from PIL import Image, ImageDraw, ImageFont, ImageOps
    except ImportError as exc:
        raise RuntimeError("服务器缺少 Pillow，无法处理图片。请先安装 requirements.txt 中的依赖。") from exc
    return Image, ImageDraw, ImageFont, ImageOps


def save_showcase_upload(image_file: UploadFile | None):
    if image_file is None or not image_file.filename:
        return ""

    suffix = Path(image_file.filename).suffix.lower()
    if suffix not in [".jpg", ".jpeg", ".png", ".webp", ".gif"]:
        raise ValueError("只支持 jpg、png、webp、gif 格式图片")

    try:
        Image, _, _, ImageOps = get_pillow_tools()
    except RuntimeError as exc:
        raise ValueError(str(exc)) from exc

    upload_dir = Path("app/static/uploads")
    upload_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{uuid.uuid4().hex}.jpg"
    target = upload_dir / filename

    try:
        image_file.file.seek(0)
        with Image.open(image_file.file) as source:
            source = ImageOps.exif_transpose(source)
            if source.mode in ("RGBA", "LA") or (source.mode == "P" and "transparency" in source.info):
                rgba = source.convert("RGBA")
                background = Image.new("RGB", rgba.size, (255, 255, 255))
                background.paste(rgba, mask=rgba.getchannel("A"))
                source = background
            else:
                source = source.convert("RGB")

            resized = ImageOps.fit(source, SHOWCASE_IMAGE_SIZE, method=Image.Resampling.LANCZOS)
            resized.save(target, format="JPEG", quality=88, optimize=True)
    except PermissionError as exc:
        raise ValueError("上传目录没有写入权限，请联系管理员修复 app/static/uploads 权限") from exc
    except Exception as exc:
        raise ValueError("图片处理失败，请确认文件是有效图片") from exc

    return f"/static/uploads/{filename}"


def load_quote_font(size: int, bold: bool = False):
    _, _, ImageFont, _ = get_pillow_tools()
    candidates = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc" if bold else "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc" if bold else "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for path in candidates:
        if path and Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def normalize_quote_rows(rows):
    if not isinstance(rows, list):
        return []

    normalized = []
    for row in rows[:80]:
        if not isinstance(row, dict):
            continue
        name = str(row.get("name") or row.get("title") or "").strip()
        if not name:
            continue
        quantity = str(row.get("quantity") or "1").strip()[:20]
        unit_price = str(row.get("price") or row.get("unit_price") or "").strip()[:30]
        quantity_value = parse_quote_number(quantity, default=1.0)
        price_value = parse_quote_number(unit_price, default=0.0)
        amount_value = quantity_value * price_value
        normalized.append({
            "name": name[:80],
            "image_url": str(row.get("image_url") or "").strip()[:500],
            "size": str(row.get("size") or "").strip()[:80],
            "quantity": quantity,
            "price": unit_price,
            "amount": format_quote_money(amount_value),
            "amount_value": amount_value,
        })
    return normalized


def parse_quote_number(value, default: float = 0.0) -> float:
    text_value = str(value or "").replace(",", "").strip()
    match = re.search(r"-?\d+(?:\.\d+)?", text_value)
    if not match:
        return default
    try:
        return float(match.group(0))
    except ValueError:
        return default


def format_quote_money(value: float) -> str:
    return f"{value:.2f}"


def resolve_static_image_path(image_url: str):
    if not image_url.startswith("/static/"):
        return None

    static_root = Path("app/static").resolve()
    candidate = (Path("app") / image_url.lstrip("/")).resolve()
    try:
        candidate.relative_to(static_root)
    except ValueError:
        return None
    if candidate.is_file():
        return candidate
    return None


def read_quote_image_bytes(image_url: str):
    if not image_url:
        return None

    local_path = resolve_static_image_path(image_url)
    if local_path:
        return local_path.read_bytes()
    return None


def open_quote_image(image_url: str):
    image_bytes = read_quote_image_bytes(image_url)
    if not image_bytes:
        return None

    Image, _, _, ImageOps = get_pillow_tools()
    try:
        with Image.open(BytesIO(image_bytes)) as source:
            source = ImageOps.exif_transpose(source).convert("RGB")
            return ImageOps.contain(source, (260, 180), method=Image.Resampling.LANCZOS)
    except Exception:
        return None


def text_width(draw, text, font):
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[2] - bbox[0]


def wrap_text(draw, text, font, max_width):
    if not text:
        return [""]

    lines = []
    current = ""
    for char in text:
        candidate = current + char
        if current and text_width(draw, candidate, font) > max_width:
            lines.append(current)
            current = char
        else:
            current = candidate
    if current:
        lines.append(current)
    return lines


def draw_wrapped_text(draw, text, xy, font, fill, max_width, line_height, max_lines=3):
    x, y = xy
    lines = wrap_text(draw, text, font, max_width)
    if len(lines) > max_lines:
        lines = lines[:max_lines]
        lines[-1] = lines[-1].rstrip("。,.， ") + "..."
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        y += line_height


def build_quote_image(rows):
    Image, ImageDraw, _, _ = get_pillow_tools()
    title_font = load_quote_font(32, bold=True)
    header_font = load_quote_font(17, bold=True)
    body_font = load_quote_font(16)
    small_font = load_quote_font(13)

    margin = 40
    table_width = 1120
    col_widths = [150, 320, 160, 90, 150, 250]
    title_height = 76
    header_height = 48
    row_height = 210
    footer_height = 70
    width = table_width + margin * 2
    height = margin + title_height + header_height + row_height * len(rows) + footer_height + margin

    image = Image.new("RGB", (width, height), "#f8fafc")
    draw = ImageDraw.Draw(image)
    draw.rectangle((margin, margin, margin + table_width, height - margin), fill="#ffffff", outline="#d1d5db")
    draw.text((margin + 20, margin + 18), "报价单", font=title_font, fill="#111827")
    draw.text((margin + table_width - 230, margin + 31), datetime.now().strftime("%Y-%m-%d %H:%M"), font=small_font, fill="#6b7280")

    headers = ["名称", "图片", "尺寸", "数量", "单价", "金额"]
    y = margin + title_height
    x = margin
    for index, header in enumerate(headers):
        draw.rectangle((x, y, x + col_widths[index], y + header_height), fill="#eef2ff", outline="#cbd5e1")
        draw.text((x + 12, y + 14), header, font=header_font, fill="#1f2937")
        x += col_widths[index]

    y += header_height
    for row in rows:
        x = margin
        for width_value in col_widths:
            draw.rectangle((x, y, x + width_value, y + row_height), fill="#ffffff", outline="#e5e7eb")
            x += width_value

        draw_wrapped_text(draw, row["name"], (margin + 12, y + 18), body_font, "#111827", col_widths[0] - 24, 24, max_lines=6)

        thumb = open_quote_image(row["image_url"])
        image_cell_x = margin + col_widths[0]
        if thumb:
            thumb_x = image_cell_x + (col_widths[1] - thumb.width) // 2
            thumb_y = y + (row_height - thumb.height) // 2
            image.paste(thumb, (thumb_x, thumb_y))
        else:
            empty_x0 = image_cell_x + 30
            empty_y0 = y + 24
            empty_x1 = image_cell_x + col_widths[1] - 30
            empty_y1 = y + row_height - 24
            draw.rectangle((empty_x0, empty_y0, empty_x1, empty_y1), fill="#f3f4f6", outline="#d1d5db")
            draw.text((empty_x0 + 76, empty_y0 + 72), "无图", font=small_font, fill="#6b7280")

        size_x = margin + col_widths[0] + col_widths[1]
        draw_wrapped_text(draw, row["size"], (size_x + 12, y + 18), body_font, "#374151", col_widths[2] - 24, 24, max_lines=4)
        qty_x = size_x + col_widths[2]
        draw_wrapped_text(draw, row["quantity"], (qty_x + 12, y + 18), body_font, "#374151", col_widths[3] - 24, 24, max_lines=2)
        unit_price_x = qty_x + col_widths[3]
        draw_wrapped_text(draw, row["price"], (unit_price_x + 12, y + 18), body_font, "#374151", col_widths[4] - 24, 24, max_lines=3)
        amount_x = unit_price_x + col_widths[4]
        draw_wrapped_text(draw, row["amount"], (amount_x + 12, y + 18), body_font, "#111827", col_widths[5] - 24, 24, max_lines=2)
        y += row_height

    total_value = sum(row.get("amount_value") or 0 for row in rows)
    draw.text((margin + 20, y + 18), "此报价单由当前页面临时生成，系统内不保存。", font=small_font, fill="#6b7280")
    total_text = f"总计：{format_quote_money(total_value)}"
    total_width = text_width(draw, total_text, title_font)
    draw.text((margin + table_width - total_width - 20, y + 14), total_text, font=title_font, fill="#dc2626")
    output = BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    return output


def build_quote_excel(rows):
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl，无法生成 Excel。请先安装 requirements.txt 中的依赖。") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "报价单"
    sheet.append(["货物名称", "图片", "尺寸", "数量", "单价", "金额"])
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["F"].width = 16

    header_fill = PatternFill("solid", fgColor="EEF2FF")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    image_refs = []

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, row in enumerate(rows, start=2):
        sheet.cell(index, 1, row["name"])
        sheet.cell(index, 3, row["size"])
        sheet.cell(index, 4, row["quantity"])
        sheet.cell(index, 5, row["price"])
        sheet.cell(index, 6, row["amount"])
        sheet.row_dimensions[index].height = 86

        for column in range(1, 7):
            cell = sheet.cell(index, column)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        thumb = open_quote_image(row["image_url"])
        if thumb:
            buffer = BytesIO()
            thumb.save(buffer, format="PNG")
            buffer.seek(0)
            excel_image = ExcelImage(buffer)
            excel_image.width = 88
            excel_image.height = 88
            sheet.add_image(excel_image, f"B{index}")
            image_refs.append(buffer)

    total_row = len(rows) + 2
    sheet.cell(total_row, 5, "总计")
    sheet.cell(total_row, 6, format_quote_money(sum(row.get("amount_value") or 0 for row in rows)))
    for column in range(1, 7):
        cell = sheet.cell(total_row, column)
        cell.border = border
        cell.font = Font(bold=True, color="111827")

    workbook._quote_image_refs = image_refs
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

# ========================
# 登录与退出
# ========================

@app.get("/login")
def login_page(request: Request):
    if is_logged_in(request):
        return RedirectResponse(url="/dashboard", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="login.html",
        context={"error": ""}
    )


@app.post("/login")
def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...)
):
    db: Session = SessionLocal()
    try:
        user = db.query(User).filter(User.username == username.strip()).first()

        if user is None or not user.is_active:
            return templates.TemplateResponse(
                request=request,
                name="login.html",
                context={"error": "用户名或密码错误"}
            )

        if not verify_password(password, user.password_hash):
            return templates.TemplateResponse(
                request=request,
                name="login.html",
                context={"error": "用户名或密码错误"}
            )

        request.session["user"] = user.username
        return RedirectResponse(url="/dashboard", status_code=303)
    finally:
        db.close()


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


# ========================
# 首页 / Dashboard
# ========================

@app.get("/")
def root():
    return RedirectResponse(url="/dashboard", status_code=303)


@app.get("/dashboard")
def dashboard(request: Request):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)

        priority_rank = get_priority_rank_expr()

        pending_orders = (
            db.query(Order)
            .filter(Order.print_status == "未打印")
            .order_by(
                priority_rank.asc(),
                Order.due_date.is_(None).asc(),
                Order.due_date.asc(),
                Order.order_no.desc()
            )
            .all()
        )

        return templates.TemplateResponse(
            request=request,
            name="dashboard.html",
            context={
                "current_user": current_user,
                "pending_orders": pending_orders
            }
        )
    finally:
        db.close()


# ========================
# 订单列表 / 搜索 / 筛选 / 排序
# ========================

@app.get("/orders")
def order_list(
    request: Request,
    keyword: str = "",
    payment_status: str = "",
    production_status: str = "",
    print_status: str = "未打印",
    date_from: str = "",
    date_to: str = "",
    sort_by: str = "priority_due"
):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)

        query = db.query(Order)

        if keyword.strip():
            kw = f"%{keyword.strip()}%"
            query = query.filter(
                or_(
                    Order.order_no.like(kw),
                    Order.customer.like(kw),
                    Order.item_name.like(kw),
                    Order.size.like(kw),
                    Order.remark.like(kw)
                )
            )

        if payment_status.strip():
            query = query.filter(Order.payment_status == payment_status.strip())

        if production_status.strip():
            query = query.filter(Order.production_status == production_status.strip())

        if print_status.strip():
            query = query.filter(Order.print_status == print_status.strip())

        parsed_from, from_error = parse_date_field(date_from, "开始日期")
        parsed_to, to_error = parse_date_field(date_to, "结束日期")
        for error in [from_error, to_error]:
            if error:
                add_flash(request, error, "error")

        if parsed_from:
            query = query.filter(Order.created_at >= datetime.combine(parsed_from, datetime.min.time()))

        if parsed_to:
            query = query.filter(Order.created_at <= datetime.combine(parsed_to, datetime.max.time()))

        if sort_by == "payment_first":
            payment_order = case(
                (Order.payment_status == "未付款", 0),
                else_=2
            )
            query = query.order_by(payment_order.asc(), Order.order_no.desc())

        elif sort_by == "order_new":
            query = query.order_by(Order.order_no.desc())

        elif sort_by == "order_old":
            query = query.order_by(Order.order_no.asc())

        elif sort_by == "customer":
            query = query.order_by(Order.customer.asc(), Order.order_no.desc())

        elif sort_by == "item":
            query = query.order_by(Order.item_name.asc(), Order.order_no.desc())

        elif sort_by == "amount_desc":
            query = query.order_by(Order.total_amount.desc(), Order.order_no.desc())

        elif sort_by == "amount_asc":
            query = query.order_by(Order.total_amount.asc(), Order.order_no.desc())

        elif sort_by == "unpaid_desc":
            query = query.order_by(Order.unpaid_amount.desc(), Order.order_no.desc())

        elif sort_by == "priority_due":
            priority_rank = get_priority_rank_expr()
            query = query.order_by(
                priority_rank.asc(),
                Order.due_date.is_(None).asc(),
                Order.due_date.asc(),
                Order.order_no.desc()
            )

        else:
            query = query.order_by(Order.order_no.desc())

        orders = query.all()

        total_amount_sum = sum(order.total_amount or 0 for order in orders)
        balance_due_sum = sum(order.unpaid_amount or 0 for order in orders)

        return templates.TemplateResponse(
            request=request,
            name="orders.html",
            context={
                "orders": orders,
                "keyword": keyword,
                "payment_status": payment_status,
                "production_status": production_status,
                "print_status": print_status,
                "date_from": date_from if not from_error else "",
                "date_to": date_to if not to_error else "",
                "sort_by": sort_by,
                "total_amount_sum": total_amount_sum,
                "balance_due_sum": balance_due_sum,
                "current_user": current_user
            }
        )
    finally:
        db.close()

@app.get("/api/suggest")
def suggest_values(
    request: Request,
    field: str,
    q: str = ""
):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return JSONResponse({"items": []}, status_code=401)

        q = q.strip()

        allowed_fields = {
            "customer": Order.customer,
            "item_name": Order.item_name,
            "size": Order.size,
        }

        if field not in allowed_fields:
            return {"items": []}

        column = allowed_fields[field]

        items = get_recent_distinct_values(db, column, q, limit=20)
        return {"items": items}
    finally:
        db.close()


@app.get("/api/recent-price")
def recent_price(
    request: Request,
    item_name: str = "",
    size: str = ""
):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return JSONResponse({"ok": False}, status_code=401)

        item_name = item_name.strip()
        size = size.strip()
        if not item_name:
            return {"ok": False}

        query = db.query(Order).filter(Order.item_name == item_name)
        if size:
            query = query.filter(Order.size == size)

        recent_order = query.order_by(Order.created_at.desc(), Order.id.desc()).first()
        if not recent_order:
            return {"ok": False}

        return {
            "ok": True,
            "unit_price": recent_order.unit_price,
            "order_no": recent_order.order_no,
        }
    finally:
        db.close()


@app.post("/api/voice-order-query-draft")
async def voice_order_query_draft(request: Request):
    redirect = require_login(request)
    if redirect:
        return JSONResponse({"ok": False, "message": "请先登录"}, status_code=401)

    payload = await request.json()
    text_value = str(payload.get("text") or "").strip()
    if not text_value:
        return JSONResponse({"ok": False, "message": "请输入语音识别后的查询内容"}, status_code=400)

    db: Session = SessionLocal()
    try:
        hotwords = build_order_hotwords(db)
        flash_model = os.getenv("DEEPSEEK_FLASH_MODEL", "deepseek-v4-flash")
        pro_model = os.getenv("DEEPSEEK_PRO_MODEL", "deepseek-v4-pro")

        raw, message = call_deepseek_query_model(flash_model, text_value, hotwords)
        source_model = flash_model
        if raw is None:
            draft, issues, confidence = heuristic_order_query(text_value, hotwords)
            source_model = "local-rules"
            if message:
                issues.append(message)
        else:
            draft, issues, confidence = normalize_order_query(raw)

        if confidence < 0.65 or issues:
            pro_raw, pro_message = call_deepseek_query_model(pro_model, text_value, hotwords)
            if pro_raw is not None:
                draft, issues, confidence = normalize_order_query(pro_raw)
                source_model = pro_model
            elif pro_message:
                issues.append(pro_message)

        return {
            "ok": True,
            "draft": draft,
            "issues": issues,
            "confidence": confidence,
            "source_model": source_model,
            "needs_review": True,
        }
    finally:
        db.close()


@app.post("/api/voice-order-draft")
async def voice_order_draft(request: Request):
    redirect = require_login(request)
    if redirect:
        return JSONResponse({"ok": False, "message": "请先登录"}, status_code=401)

    payload = await request.json()
    text_value = str(payload.get("text") or "").strip()
    if not text_value:
        return JSONResponse({"ok": False, "message": "请输入语音识别后的文字"}, status_code=400)

    db: Session = SessionLocal()
    try:
        hotwords = build_order_hotwords(db)
        flash_model = os.getenv("DEEPSEEK_FLASH_MODEL", "deepseek-v4-flash")
        pro_model = os.getenv("DEEPSEEK_PRO_MODEL", "deepseek-v4-pro")

        raw, message = call_deepseek_model(flash_model, text_value, hotwords)
        source_model = flash_model
        if raw is None:
            draft, issues = heuristic_order_draft(text_value, hotwords)
            draft = fill_recent_unit_price(db, draft)
            draft, issues = normalize_order_draft(draft)
            result = {
                "orders": [{"draft": draft, "issues": issues}],
                "issues": issues[:],
                "confidence": 0.45,
            }
            source_model = "local-rules"
            if message:
                result["issues"].append(message)
        else:
            result = normalize_order_result(raw, text_value, db, hotwords)

        conflict_words = ["两个订单", "多个订单", "分别", "另外", "再来一单", "还有一单"]
        needs_pro = (
            len(result["orders"]) > 1
            or len(result["issues"]) >= 2
            or any(word in text_value for word in conflict_words)
            or result["confidence"] < 0.65
        )
        if needs_pro and source_model != "local-rules":
            pro_raw, pro_message = call_deepseek_model(pro_model, text_value, hotwords)
            if pro_raw is not None:
                result = normalize_order_result(pro_raw, text_value, db, hotwords)
                source_model = pro_model
            elif pro_message:
                result["issues"].append(pro_message)

        first_draft = result["orders"][0]["draft"] if result["orders"] else {}
        return {
            "ok": True,
            "draft": first_draft,
            "orders": result["orders"],
            "issues": result["issues"],
            "confidence": result["confidence"],
            "source_model": source_model,
            "needs_review": bool(result["issues"]) or len(result["orders"]) > 1,
            "hotwords": hotwords,
        }
    finally:
        db.close()


# ========================
# 新建订单
# ========================

@app.get("/orders/new")
def order_new(request: Request):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        next_order_no = generate_order_no(db)

        customer_options = get_recent_distinct_values(db, Order.customer, limit=200)
        item_options = get_recent_distinct_values(db, Order.item_name, limit=200)
        size_options = get_recent_distinct_values(db, Order.size, limit=200)

        return templates.TemplateResponse(
            request=request,
            name="order_new.html",
            context={
                "next_order_no": next_order_no,
                "customer_options": customer_options,
                "item_options": item_options,
                "size_options": size_options,
                "current_user": current_user,
                "form_data": {}
            }
        )
    finally:
        db.close()


@app.post("/orders")
def create_order(
    request: Request,
    customer: str = Form(...),
    phone: str = Form(""),
    item_name: str = Form(...),
    size: str = Form(""),
    quantity: int = Form(...),
    unit_price: float = Form(...),
    paid_amount: float = Form(0.0),
    priority_color: str = Form("灰色"),
    due_date: str = Form(""),
    remark: str = Form("")
):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        errors, parsed_due_date = order_form_error(
            customer,
            item_name,
            quantity,
            unit_price,
            paid_amount,
            priority_color,
            due_date,
        )
        if errors:
            current_user = get_current_user(request, db)
            return templates.TemplateResponse(
                request=request,
                name="order_new.html",
                context={
                    "next_order_no": generate_order_no(db),
                    "customer_options": get_recent_distinct_values(db, Order.customer, limit=200),
                    "item_options": get_recent_distinct_values(db, Order.item_name, limit=200),
                    "size_options": get_recent_distinct_values(db, Order.size, limit=200),
                    "current_user": current_user,
                    "form_data": {
                        "customer": customer,
                        "phone": phone,
                        "item_name": item_name,
                        "size": size,
                        "quantity": quantity,
                        "unit_price": unit_price,
                        "paid_amount": paid_amount,
                        "priority_color": priority_color,
                        "due_date": due_date,
                        "remark": remark,
                    },
                    "errors": errors,
                },
                status_code=400,
            )

        order_no = generate_order_no(db)
        total_amount = quantity * unit_price
        unpaid_amount, payment_status_value = calc_payment_status(total_amount, paid_amount)

        new_order = Order(
            order_no=order_no,
            customer=customer,
            phone=phone,
            item_name=item_name,
            size=size,
            quantity=quantity,
            unit_price=unit_price,
            total_amount=total_amount,
            paid_amount=paid_amount,
            unpaid_amount=unpaid_amount,
            payment_status=payment_status_value,
            production_status="未投产",
            print_status="未打印",
            priority_color=priority_color,
            due_date=parsed_due_date,
            remark=remark
        )

        db.add(new_order)
        db.commit()

        add_flash(request, f"订单 {order_no} 已创建", "success")
        return RedirectResponse(url="/orders", status_code=303)
    finally:
        db.close()


# ========================
# 订单详情 / 编辑
# ========================

@app.get("/orders/{order_id}")
def order_detail(request: Request, order_id: int):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        order = db.query(Order).filter(Order.id == order_id).first()

        if order is None:
            return templates.TemplateResponse(
                request=request,
                name="not_found.html",
                context={"message": f"订单 ID {order_id} 不存在", "current_user": current_user},
                status_code=404
            )

        return templates.TemplateResponse(
            request=request,
            name="order_detail.html",
            context={
                "order": order,
                "current_user": current_user,
                "logs": (
                    db.query(OperationLog)
                    .filter(OperationLog.target_type == "order", OperationLog.target_id == order.id)
                    .order_by(OperationLog.id.desc())
                    .limit(12)
                    .all()
                )
            }
        )
    finally:
        db.close()


@app.get("/orders/{order_id}/print-preview")
def print_preview_page(request: Request, order_id: int, print_template: str = "auto"):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        order = db.query(Order).filter(Order.id == order_id).first()
        if order is None:
            return templates.TemplateResponse(
                request=request,
                name="not_found.html",
                context={"message": f"订单 ID {order_id} 不存在", "current_user": current_user},
                status_code=404
            )

        template_key = resolve_print_template(db, order, print_template)
        return templates.TemplateResponse(
            request=request,
            name="print_preview.html",
            context={
                "order": order,
                "current_user": current_user,
                "template_key": template_key,
                "requested_template": str(print_template or "auto"),
                "template_label": PRINT_TEMPLATES[template_key]["label"],
                "preview_text": build_print_text(order, template_key),
            }
        )
    finally:
        db.close()


@app.get("/orders/{order_id}/edit")
def edit_order_page(request: Request, order_id: int):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        order = db.query(Order).filter(Order.id == order_id).first()

        if order is None:
            return templates.TemplateResponse(
                request=request,
                name="not_found.html",
                context={"message": f"订单 ID {order_id} 不存在", "current_user": current_user},
                status_code=404
            )

        customer_options = [
            row[0] for row in db.query(Order.customer).distinct().order_by(Order.customer.asc()).all() if row[0]
        ]
        item_options = [
            row[0] for row in db.query(Order.item_name).distinct().order_by(Order.item_name.asc()).all() if row[0]
        ]
        size_options = [
            row[0] for row in db.query(Order.size).distinct().order_by(Order.size.asc()).all() if row[0]
        ]

        return templates.TemplateResponse(
            request=request,
            name="order_edit.html",
            context={
                "order": order,
                "customer_options": customer_options,
                "item_options": item_options,
                "size_options": size_options,
                "current_user": current_user,
                "errors": []
            }
        )
    finally:
        db.close()


@app.post("/orders/{order_id}/edit")
def edit_order_submit(
    request: Request,
    order_id: int,
    customer: str = Form(...),
    phone: str = Form(""),
    item_name: str = Form(...),
    size: str = Form(""),
    quantity: int = Form(...),
    unit_price: float = Form(...),
    paid_amount: float = Form(0.0),
    priority_color: str = Form("灰色"),
    due_date: str = Form(""),
    remark: str = Form("")
):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        order = db.query(Order).filter(Order.id == order_id).first()
        if order is None:
            return RedirectResponse(url="/orders", status_code=303)

        errors, parsed_due_date = order_form_error(
            customer,
            item_name,
            quantity,
            unit_price,
            paid_amount,
            priority_color,
            due_date,
        )
        if errors:
            current_user = get_current_user(request, db)
            return templates.TemplateResponse(
                request=request,
                name="order_edit.html",
                context={
                    "order": order,
                    "customer_options": get_recent_distinct_values(db, Order.customer, limit=200),
                    "item_options": get_recent_distinct_values(db, Order.item_name, limit=200),
                    "size_options": get_recent_distinct_values(db, Order.size, limit=200),
                    "current_user": current_user,
                    "errors": errors,
                    "form_data": {
                        "customer": customer,
                        "phone": phone,
                        "item_name": item_name,
                        "size": size,
                        "quantity": quantity,
                        "unit_price": unit_price,
                        "paid_amount": paid_amount,
                        "priority_color": priority_color,
                        "due_date": due_date,
                        "remark": remark,
                    },
                },
                status_code=400,
            )

        total_amount = quantity * unit_price
        unpaid_amount, payment_status_value = calc_payment_status(total_amount, paid_amount, order.payment_status)

        order.customer = customer
        order.phone = phone
        order.item_name = item_name
        order.size = size
        order.quantity = quantity
        order.unit_price = unit_price
        order.total_amount = total_amount
        order.paid_amount = paid_amount
        order.unpaid_amount = unpaid_amount
        order.payment_status = payment_status_value
        order.priority_color = priority_color
        order.due_date = parsed_due_date
        order.remark = remark


        current_user = get_current_user(request, db)

        log_operation(
            db=db,
            target_type="order",
            target_id=order.id,
            action="edit_order",
            field_name="multiple",
            old_value="订单被编辑",
            new_value=f"customer={customer}, item_name={item_name}, size={size}, quantity={quantity}, unit_price={unit_price}, paid_amount={paid_amount}",
            operator=current_user.username if current_user else ""
        )
        db.commit()

        add_flash(request, f"订单 {order.order_no} 已保存", "success")
        return RedirectResponse(url=f"/orders/{order_id}", status_code=303)
    finally:
        db.close()


# ========================
# 订单状态操作
# ========================

@app.post("/orders/{order_id}/print")
def mark_order_printed(
    request: Request,
    order_id: int,
    print_template: str = Form("auto"),
    return_to: str = Form("")
):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)

        order = db.query(Order).filter(Order.id == order_id).first()
        if order is None:
            add_flash(request, "订单不存在，无法加入打印队列", "error")
            return RedirectResponse(url="/orders", status_code=303)

        queued, skipped = queue_print_jobs(
            db,
            [order],
            operator=current_user.username if current_user else "",
            print_template=print_template
        )
        db.commit()
        if queued:
            add_flash(request, f"订单 {order.order_no} 已加入打印队列", "success")
        elif skipped:
            add_flash(request, f"订单 {order.order_no} 已打印或已在打印队列中", "warning")

        return RedirectResponse(url=safe_redirect_path(return_to, f"/orders/{order_id}"), status_code=303)
    finally:
        db.close()


@app.post("/orders/batch-print")
def batch_print_orders(
    request: Request,
    order_ids: list[int] = Form(default=[]),
    print_template: str = Form("auto"),
    return_to: str = Form("/orders")
):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        if order_ids:
            orders = (
                db.query(Order)
                .filter(Order.id.in_(order_ids))
                .order_by(Order.order_no.asc())
                .all()
            )
            queued, skipped = queue_print_jobs(
                db,
                orders,
                operator=current_user.username if current_user else "",
                print_template=print_template
            )
            db.commit()
            add_flash(
                request,
                f"已加入 {len(queued)} 个打印任务"
                + (f"，跳过 {len(skipped)} 个已打印或队列中订单" if skipped else ""),
                "success" if queued else "warning",
            )
        else:
            add_flash(request, "请先选择要发送打印的订单", "warning")

        return RedirectResponse(url=safe_redirect_path(return_to, "/orders"), status_code=303)
    finally:
        db.close()


@app.post("/api/print-client/next")
async def print_client_next(request: Request):
    if not verify_print_client(request):
        return JSONResponse({"ok": False, "message": "打印客户端令牌无效或服务器未配置 PRINT_CLIENT_TOKEN"}, status_code=401)

    payload = await request.json()
    client_id = str(payload.get("client_id") or "win7-print-client").strip()

    db: Session = SessionLocal()
    try:
        stale_before = datetime.utcnow() - timedelta(minutes=10)
        job = (
            db.query(PrintJob)
            .join(Order, Order.id == PrintJob.order_id)
            .filter(
                Order.print_status == "未打印",
                or_(
                    PrintJob.status == "pending",
                    (PrintJob.status == "printing") & (PrintJob.claimed_at < stale_before)
                )
            )
            .order_by(PrintJob.id.asc())
            .first()
        )

        if job is None:
            return {"ok": True, "has_job": False}

        order = db.query(Order).filter(Order.id == job.order_id).first()
        if order is None:
            job.status = "failed"
            job.error_message = "订单不存在"
            db.commit()
            return {"ok": True, "has_job": False}

        job.status = "printing"
        job.client_id = client_id
        job.attempts = (job.attempts or 0) + 1
        job.claimed_at = datetime.utcnow()
        db.commit()

        return {
            "ok": True,
            "has_job": True,
            "job": {
                "id": job.id,
                "order_id": order.id,
                "order_no": order.order_no,
                "print_template": normalize_print_template(job.print_template),
            },
            "payload": {
                "format": "text",
                "template": normalize_print_template(job.print_template),
                "template_label": PRINT_TEMPLATES[normalize_print_template(job.print_template)]["label"],
                "text": build_print_text(order, job.print_template),
            },
        }
    finally:
        db.close()


@app.post("/api/print-client/report")
async def print_client_report(request: Request):
    if not verify_print_client(request):
        return JSONResponse({"ok": False, "message": "打印客户端令牌无效或服务器未配置 PRINT_CLIENT_TOKEN"}, status_code=401)

    payload = await request.json()
    job_id = int(payload.get("job_id") or 0)
    success = bool(payload.get("success"))
    error_message = str(payload.get("error") or "").strip()
    client_id = str(payload.get("client_id") or "win7-print-client").strip()

    db: Session = SessionLocal()
    try:
        job = db.query(PrintJob).filter(PrintJob.id == job_id).first()
        if job is None:
            return JSONResponse({"ok": False, "message": "打印任务不存在"}, status_code=404)

        order = db.query(Order).filter(Order.id == job.order_id).first()
        if order is None:
            job.status = "failed"
            job.error_message = "订单不存在"
            db.commit()
            return JSONResponse({"ok": False, "message": "订单不存在"}, status_code=404)

        job.client_id = client_id
        if success:
            old_value = order.print_status
            order.print_status = "已打印"
            job.status = "done"
            job.error_message = ""
            job.printed_at = datetime.utcnow()

            if old_value != order.print_status:
                log_operation(
                    db=db,
                    target_type="order",
                    target_id=order.id,
                    action="mark_printed",
                    field_name="print_status",
                    old_value=old_value,
                    new_value=order.print_status,
                    operator=client_id
                )
        else:
            job.status = "failed"
            job.error_message = error_message[:1000] or "打印失败"
            log_operation(
                db=db,
                target_type="order",
                target_id=order.id,
                action="print_failed",
                field_name="print_job",
                old_value=f"job_id={job.id}",
                new_value=job.error_message,
                operator=client_id
            )

        db.commit()
        return {"ok": True}
    finally:
        db.close()


@app.get("/print-settings")
def print_settings_page(request: Request):
    db: Session = SessionLocal()
    try:
        redirect = require_admin(request, db)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        customers = (
            db.query(Order.customer)
            .filter(Order.customer != "")
            .distinct()
            .order_by(Order.customer.asc())
            .limit(300)
            .all()
        )
        rules = (
            db.query(PrintTemplateRule)
            .order_by(PrintTemplateRule.customer_name.asc())
            .all()
        )
        return templates.TemplateResponse(
            request=request,
            name="print_settings.html",
            context={
                "current_user": current_user,
                "default_print_template": get_default_print_template(db),
                "rules": rules,
                "customers": [row[0] for row in customers if row[0]],
            }
        )
    finally:
        db.close()


@app.post("/print-settings/default")
def update_default_print_template(
    request: Request,
    default_print_template: str = Form("delivery")
):
    db: Session = SessionLocal()
    try:
        redirect = require_admin(request, db)
        if redirect:
            return redirect

        template_key = normalize_print_template(default_print_template)
        set_app_setting(db, "default_print_template", template_key)
        db.commit()
        add_flash(request, f"默认打印模板已设置为：{print_template_label(template_key)}", "success")
        return RedirectResponse(url="/print-settings", status_code=303)
    finally:
        db.close()


@app.post("/print-settings/rules")
def save_print_template_rule(
    request: Request,
    customer_name: str = Form(...),
    print_template: str = Form("delivery")
):
    db: Session = SessionLocal()
    try:
        redirect = require_admin(request, db)
        if redirect:
            return redirect

        customer_name = str(customer_name or "").strip()
        if not customer_name:
            add_flash(request, "客户名称不能为空", "error")
            return RedirectResponse(url="/print-settings", status_code=303)

        template_key = normalize_print_template(print_template)
        rule = db.query(PrintTemplateRule).filter(PrintTemplateRule.customer_name == customer_name).first()
        if rule is None:
            rule = PrintTemplateRule(
                customer_name=customer_name,
                print_template=template_key,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            )
            db.add(rule)
        else:
            rule.print_template = template_key
            rule.updated_at = datetime.utcnow()
        db.commit()
        add_flash(request, f"{customer_name} 的打印模板已设置为：{print_template_label(template_key)}", "success")
        return RedirectResponse(url="/print-settings", status_code=303)
    finally:
        db.close()


@app.post("/print-settings/rules/{rule_id}/delete")
def delete_print_template_rule(request: Request, rule_id: int):
    db: Session = SessionLocal()
    try:
        redirect = require_admin(request, db)
        if redirect:
            return redirect

        rule = db.query(PrintTemplateRule).filter(PrintTemplateRule.id == rule_id).first()
        if rule is not None:
            db.delete(rule)
            db.commit()
            add_flash(request, f"已删除 {rule.customer_name} 的专属模板规则", "success")
        else:
            add_flash(request, "规则不存在", "warning")
        return RedirectResponse(url="/print-settings", status_code=303)
    finally:
        db.close()


@app.get("/print-jobs")
@app.get("/print_jobs")
@app.get("/print-queue")
@app.get("/print_queue")
def print_jobs_page(request: Request, status: str = ""):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        allowed_statuses = ["", "pending", "printing", "failed", "done"]
        if status not in allowed_statuses:
            status = ""

        query = (
            db.query(PrintJob, Order)
            .join(Order, Order.id == PrintJob.order_id)
        )
        if status:
            query = query.filter(PrintJob.status == status)

        rows = (
            query
            .order_by(PrintJob.id.desc())
            .limit(200)
            .all()
        )
        jobs = [
            {
                "job": job,
                "order": order,
                "status_label": get_print_status_label(job.status),
            }
            for job, order in rows
        ]

        status_counts = {
            item_status: db.query(PrintJob).filter(PrintJob.status == item_status).count()
            for item_status in ["pending", "printing", "failed", "done"]
        }
        candidate_orders = (
            db.query(Order)
            .filter(Order.print_status == "未打印")
            .order_by(Order.order_no.desc())
            .limit(100)
            .all()
        )
        available_orders = [order for order in candidate_orders if get_active_print_job(db, order.id) is None]

        return templates.TemplateResponse(
            request=request,
            name="print_jobs.html",
            context={
                "jobs": jobs,
                "status": status,
                "status_counts": status_counts,
                "available_orders": available_orders,
                "current_user": current_user,
                "default_print_template": get_default_print_template(db),
                "status_labels": {
                    "pending": "等待打印",
                    "printing": "打印中",
                    "failed": "打印失败",
                    "done": "已完成",
                },
            }
        )
    finally:
        db.close()


@app.post("/print-jobs/add-order")
def add_order_to_print_queue(
    request: Request,
    order_id: int = Form(...),
    print_template: str = Form("auto")
):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        order = db.query(Order).filter(Order.id == order_id).first()
        if order is not None:
            queued, skipped = queue_print_jobs(
                db,
                [order],
                operator=current_user.username if current_user else "",
                print_template=print_template
            )
            db.commit()
            if queued:
                add_flash(request, f"订单 {order.order_no} 已加入打印队列", "success")
            elif skipped:
                add_flash(request, f"订单 {order.order_no} 已打印或已在打印队列中", "warning")
        else:
            add_flash(request, "订单不存在，无法加入打印队列", "error")

        return RedirectResponse(url="/print-jobs?status=pending", status_code=303)
    finally:
        db.close()


@app.post("/print-jobs/{job_id}/retry")
def retry_print_job(request: Request, job_id: int):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        job = db.query(PrintJob).filter(PrintJob.id == job_id).first()
        if job is not None:
            order = db.query(Order).filter(Order.id == job.order_id).first()
            if order is not None:
                order.print_status = "未打印"
                job.status = "pending"
                job.client_id = ""
                job.error_message = ""
                job.claimed_at = None
                job.printed_at = None

                log_operation(
                    db=db,
                    target_type="order",
                    target_id=order.id,
                    action="retry_print",
                    field_name="print_job",
                    old_value=f"job_id={job.id}",
                    new_value="pending",
                    operator=current_user.username if current_user else ""
                )

                db.commit()
                add_flash(request, f"打印任务 #{job.id} 已重新加入队列", "success")
            else:
                add_flash(request, "打印任务关联的订单不存在", "error")
        else:
            add_flash(request, "打印任务不存在", "error")

        return RedirectResponse(url="/print-jobs?status=failed", status_code=303)
    finally:
        db.close()


@app.post("/orders/{order_id}/paid")
def mark_order_paid(
    request: Request,
    order_id: int,
    return_to: str = Form("")
):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        order = db.query(Order).filter(Order.id == order_id).first()
        if order is not None:
            old_value = json.dumps(payment_snapshot(order), ensure_ascii=False)
            order.payment_status = "已付款"
            order.paid_amount = float(order.total_amount or 0)
            order.unpaid_amount = 0.0
            new_value = json.dumps(payment_snapshot(order), ensure_ascii=False)

            log_operation(
                db=db,
                target_type="order",
                target_id=order.id,
                action="mark_paid",
                field_name="payment_status",
                old_value=old_value,
                new_value=new_value,
                operator=current_user.username if current_user else ""
            )

            db.commit()
            add_flash(request, f"订单 {order.order_no} 已标记结清，已收金额已同步为总金额", "success")
        else:
            add_flash(request, "订单不存在，无法标记付款", "error")

        return RedirectResponse(url=safe_redirect_path(return_to, f"/orders/{order_id}"), status_code=303)
    finally:
        db.close()


@app.post("/orders/{order_id}/unpaid")
def mark_order_unpaid(
    request: Request,
    order_id: int,
    return_to: str = Form("")
):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        order = db.query(Order).filter(Order.id == order_id).first()
        if order is not None:
            old_value = json.dumps(payment_snapshot(order), ensure_ascii=False)
            order.payment_status = "未付款"
            order.unpaid_amount = max(float(order.total_amount or 0) - float(order.paid_amount or 0), 0.0)
            if order.unpaid_amount <= 0:
                order.paid_amount = 0.0
                order.unpaid_amount = float(order.total_amount or 0)
            new_value = json.dumps(payment_snapshot(order), ensure_ascii=False)

            log_operation(
                db=db,
                target_type="order",
                target_id=order.id,
                action="mark_unpaid",
                field_name="payment_status",
                old_value=old_value,
                new_value=new_value,
                operator=current_user.username if current_user else ""
            )

            db.commit()
            add_flash(request, f"订单 {order.order_no} 已改回未付款", "success")
        else:
            add_flash(request, "订单不存在，无法改回未付款", "error")

        return RedirectResponse(url=safe_redirect_path(return_to, f"/orders/{order_id}"), status_code=303)
    finally:
        db.close()


@app.post("/orders/{order_id}/production")
def mark_order_production(request: Request, order_id: int):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)

        order = db.query(Order).filter(Order.id == order_id).first()
        if order is not None:
            old_value = order.production_status
            order.production_status = "已投产"

            log_operation(
                db=db,
                target_type="order",
                target_id=order.id,
                action="mark_production",
                field_name="production_status",
                old_value=old_value,
                new_value=order.production_status,
                operator=current_user.username if current_user else ""
            )

            db.commit()

        return RedirectResponse(url=f"/orders/{order_id}", status_code=303)
    finally:
        db.close()

@app.post("/orders/{order_id}/complete")
def mark_order_complete(request: Request, order_id: int):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)

        order = db.query(Order).filter(Order.id == order_id).first()
        if order is not None:
            old_value = order.production_status
            order.production_status = "已完成"

            log_operation(
                db=db,
                target_type="order",
                target_id=order.id,
                action="mark_complete",
                field_name="production_status",
                old_value=old_value,
                new_value=order.production_status,
                operator=current_user.username if current_user else ""
            )

            db.commit()

        return RedirectResponse(url=f"/orders/{order_id}", status_code=303)
    finally:
        db.close()


@app.post("/orders/{order_id}/delete")
def delete_order(request: Request, order_id: int):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        order = db.query(Order).filter(Order.id == order_id).first()

        if order is not None:
            active_job = get_active_print_job(db, order.id)
            if active_job is not None:
                add_flash(request, f"订单 {order.order_no} 已在打印队列中，先处理打印任务后再删除", "error")
                return RedirectResponse(url=f"/orders/{order_id}", status_code=303)

            log_operation(
                db=db,
                target_type="order",
                target_id=order.id,
                action="delete_order",
                field_name="order",
                old_value=f"order_no={order.order_no}",
                new_value="deleted",
                operator=current_user.username if current_user else ""
            )

            db.query(PrintJob).filter(PrintJob.order_id == order.id).delete(synchronize_session=False)
            db.delete(order)
            db.commit()
            add_flash(request, f"订单 {order.order_no} 已删除", "success")
        else:
            add_flash(request, "订单不存在或已被删除", "warning")

        return RedirectResponse(url="/orders", status_code=303)
    finally:
        db.close()

@app.post("/orders/{order_id}/undo")
def undo_last_order_action(request: Request, order_id: int):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        order = db.query(Order).filter(Order.id == order_id).first()

        if order is None:
            add_flash(request, "订单不存在，无法撤回", "error")
            return RedirectResponse(url="/orders", status_code=303)

        log = get_latest_reversible_log(db, order_id)
        if log is None:
            add_flash(request, "没有可撤回的状态操作", "warning")
            return RedirectResponse(url=f"/orders/{order_id}", status_code=303)

        if log.field_name == "print_status":
            current_new_value = order.print_status
            order.print_status = log.old_value

            log_operation(
                db=db,
                target_type="order",
                target_id=order.id,
                action="undo_print_status",
                field_name="print_status",
                old_value=current_new_value,
                new_value=order.print_status,
                operator=current_user.username if current_user else ""
            )

        elif log.field_name == "production_status":
            current_new_value = order.production_status
            order.production_status = log.old_value

            log_operation(
                db=db,
                target_type="order",
                target_id=order.id,
                action="undo_production_status",
                field_name="production_status",
                old_value=current_new_value,
                new_value=order.production_status,
                operator=current_user.username if current_user else ""
            )

        elif log.field_name == "payment_status":
            current_new_value = order.payment_status
            old_snapshot = parse_payment_snapshot(log.old_value)
            order.payment_status = old_snapshot.get("payment_status") or "未付款"
            if "paid_amount" in old_snapshot:
                order.paid_amount = float(old_snapshot.get("paid_amount") or 0)
                order.unpaid_amount = float(old_snapshot.get("unpaid_amount") or 0)
            elif order.payment_status == "已付款":
                order.paid_amount = float(order.total_amount or 0)
                order.unpaid_amount = 0.0
            else:
                order.unpaid_amount, order.payment_status = calc_payment_status(
                    order.total_amount,
                    order.paid_amount,
                    "未付款"
                )

            log_operation(
                db=db,
                target_type="order",
                target_id=order.id,
                action="undo_payment_status",
                field_name="payment_status",
                old_value=current_new_value,
                new_value=order.payment_status,
                operator=current_user.username if current_user else ""
            )

        db.commit()

        add_flash(request, f"已撤回：{action_label(log.action)}", "success")
        return RedirectResponse(url=f"/orders/{order_id}", status_code=303)
    finally:
        db.close()


# ========================
# 展示页（游客可访问）
# ========================

@app.get("/showcase")
def showcase_public(request: Request, category: str = "", q: str = ""):
    db: Session = SessionLocal()
    try:
        query = db.query(ShowcaseItem).filter(ShowcaseItem.is_visible == True)
        if category.strip():
            query = query.filter(ShowcaseItem.category == category.strip())
        if q.strip():
            query = query.filter(ShowcaseItem.title.like(f"%{q.strip()}%"))

        items = query.order_by(ShowcaseItem.category.asc(), ShowcaseItem.item_code.asc(), ShowcaseItem.id.asc()).all()
        categories = (
            db.query(ShowcaseItem.category)
            .filter(ShowcaseItem.is_visible == True)
            .distinct()
            .order_by(ShowcaseItem.category.asc())
            .all()
        )

        return templates.TemplateResponse(
            request=request,
            name="showcase.html",
            context={
                "items": items,
                "categories": [row[0] for row in categories if row[0]],
                "current_category": category,
                "current_query": q.strip()
            }
        )
    finally:
        db.close()


# ========================
# 展示管理（需登录）
# ========================

@app.get("/showcase/manage")
def showcase_manage(request: Request, category: str = ""):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        query = db.query(ShowcaseItem)
        if category.strip():
            query = query.filter(ShowcaseItem.category == category.strip())
        items = query.order_by(ShowcaseItem.category.asc(), ShowcaseItem.item_code.asc(), ShowcaseItem.id.asc()).all()
        categories = get_showcase_category_options(db)

        return templates.TemplateResponse(
            request=request,
            name="showcase_manage.html",
            context={
                "items": items,
                "current_user": current_user,
                "categories": categories,
                "current_category": category.strip(),
            }
        )
    finally:
        db.close()


@app.get("/showcase/manage/new")
def showcase_new_page(request: Request):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        categories = get_showcase_category_options(db)

        return templates.TemplateResponse(
            request=request,
            name="showcase_new.html",
            context={"error": "", "current_user": current_user, "categories": categories}
        )
    finally:
        db.close()


@app.post("/showcase/manage/new")
def showcase_new_submit(
    request: Request,
    title: str = Form(...),
    category: str = Form(""),
    image_file: UploadFile = File(None),
    description: str = Form(""),
    is_visible: str = Form("true")
):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return redirect

        try:
            uploaded_url = save_showcase_upload(image_file)
        except ValueError as exc:
            current_user = get_current_user(request, db)
            categories = get_showcase_category_options(db)
            return templates.TemplateResponse(
                request=request,
                name="showcase_new.html",
                context={"error": str(exc), "current_user": current_user, "categories": categories},
                status_code=400
            )

        if not title.strip():
            current_user = get_current_user(request, db)
            categories = get_showcase_category_options(db)
            return templates.TemplateResponse(
                request=request,
                name="showcase_new.html",
                context={"error": "标题不能为空", "current_user": current_user, "categories": categories},
                status_code=400
            )

        category_value = category.strip() or "未分类"
        item = ShowcaseItem(
            title=title.strip(),
            item_code=generate_showcase_item_code(db, category_value),
            category=category_value,
            image_url=uploaded_url,
            description=description.strip(),
            is_visible=(is_visible == "true")
        )

        db.add(item)
        db.commit()

        add_flash(request, f"货物 {item.title} 已保存", "success")
        return RedirectResponse(url="/showcase/manage", status_code=303)
    finally:
        db.close()


@app.post("/showcase/manage/delete")
async def showcase_delete_items(request: Request):
    db: Session = SessionLocal()
    try:
        redirect = require_login(request)
        if redirect:
            return JSONResponse({"error": "请先登录"}, status_code=401)

        payload = await request.json()
        raw_ids = payload.get("ids") if isinstance(payload, dict) else []
        item_ids = []
        for raw_id in raw_ids if isinstance(raw_ids, list) else []:
            try:
                item_ids.append(int(raw_id))
            except (TypeError, ValueError):
                continue

        item_ids = list(dict.fromkeys(item_ids))[:100]
        if not item_ids:
            return JSONResponse({"error": "请至少选择一个货物"}, status_code=400)

        current_user = get_current_user(request, db)
        items = db.query(ShowcaseItem).filter(ShowcaseItem.id.in_(item_ids)).all()
        for item in items:
            log_operation(
                db=db,
                target_type="showcase",
                target_id=item.id,
                action="delete_showcase",
                field_name="item",
                old_value=item.title,
                new_value="deleted",
                operator=current_user.username if current_user else ""
            )
            db.delete(item)

        db.commit()
        add_flash(request, f"已删除 {len(items)} 个货物", "success")
        return {"ok": True, "deleted": len(items)}
    finally:
        db.close()


@app.post("/showcase/quotation/image")
async def showcase_quotation_image(request: Request):
    redirect = require_login(request)
    if redirect:
        return JSONResponse({"error": "请先登录"}, status_code=401)
    payload = await request.json()
    rows = normalize_quote_rows(payload.get("rows") if isinstance(payload, dict) else payload)
    if not rows:
        return JSONResponse({"error": "请至少选择一个货物"}, status_code=400)

    try:
        output = build_quote_image(rows)
    except RuntimeError as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)

    filename = f"quotation-{datetime.now().strftime('%Y%m%d%H%M%S')}.png"
    return StreamingResponse(
        output,
        media_type="image/png",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.post("/showcase/quotation/excel")
async def showcase_quotation_excel(request: Request):
    redirect = require_login(request)
    if redirect:
        return JSONResponse({"error": "请先登录"}, status_code=401)
    payload = await request.json()
    rows = normalize_quote_rows(payload.get("rows") if isinstance(payload, dict) else payload)
    if not rows:
        return JSONResponse({"error": "请至少选择一个货物"}, status_code=400)

    try:
        output = build_quote_excel(rows)
    except RuntimeError as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)

    filename = f"quotation-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ========================
# 用户管理（保留）
# ========================

@app.get("/users")
def user_list(request: Request):
    db: Session = SessionLocal()
    try:
        redirect = require_admin(request, db)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        users = db.query(User).order_by(User.id.asc()).all()

        return templates.TemplateResponse(
            request=request,
            name="users.html",
            context={"users": users, "current_user": current_user}
        )
    finally:
        db.close()


@app.get("/users/new")
def user_new_page(request: Request):
    db: Session = SessionLocal()
    try:
        redirect = require_admin(request, db)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        return templates.TemplateResponse(
            request=request,
            name="user_new.html",
            context={"error": "", "current_user": current_user}
        )
    finally:
        db.close()


@app.post("/users/new")
def user_new_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    confirm_password: str = Form(...),
    is_admin: str = Form("false")
):
    db: Session = SessionLocal()
    try:
        redirect = require_admin(request, db)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        username = username.strip()

        if not username:
            return templates.TemplateResponse(
                request=request,
                name="user_new.html",
                context={"error": "用户名不能为空", "current_user": current_user},
                status_code=400
            )

        if password != confirm_password:
            return templates.TemplateResponse(
                request=request,
                name="user_new.html",
                context={"error": "两次输入的密码不一致", "current_user": current_user},
                status_code=400
            )

        existing_user = db.query(User).filter(User.username == username).first()
        if existing_user is not None:
            return templates.TemplateResponse(
                request=request,
                name="user_new.html",
                context={"error": "用户名已存在", "current_user": current_user},
                status_code=400
            )

        user = User(
            username=username,
            password_hash=hash_password(password),
            is_active=True,
            is_admin=(is_admin == "true")
        )
        db.add(user)
        db.commit()

        add_flash(request, f"用户 {username} 已创建", "success")
        return RedirectResponse(url="/users", status_code=303)
    finally:
        db.close()


@app.get("/users/{user_id}/edit")
def user_edit_page(request: Request, user_id: int):
    db: Session = SessionLocal()
    try:
        redirect = require_admin(request, db)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        user_obj = db.query(User).filter(User.id == user_id).first()
        if user_obj is None:
            return templates.TemplateResponse(
                request=request,
                name="not_found.html",
                context={"message": f"用户 ID {user_id} 不存在", "current_user": current_user},
                status_code=404
            )

        return templates.TemplateResponse(
            request=request,
            name="user_edit.html",
            context={"user_obj": user_obj, "error": "", "current_user": current_user}
        )
    finally:
        db.close()


@app.post("/users/{user_id}/edit")
def user_edit_submit(
    request: Request,
    user_id: int,
    username: str = Form(...),
    is_active: str = Form("true"),
    is_admin: str = Form("false"),
    new_password: str = Form(""),
    confirm_password: str = Form("")
):
    db: Session = SessionLocal()
    try:
        redirect = require_admin(request, db)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        user_obj = db.query(User).filter(User.id == user_id).first()
        if user_obj is None:
            return RedirectResponse(url="/users", status_code=303)

        username = username.strip()
        if not username:
            return templates.TemplateResponse(
                request=request,
                name="user_edit.html",
                context={"user_obj": user_obj, "error": "用户名不能为空", "current_user": current_user},
                status_code=400
            )

        duplicate = (
            db.query(User)
            .filter(User.username == username, User.id != user_id)
            .first()
        )
        if duplicate is not None:
            return templates.TemplateResponse(
                request=request,
                name="user_edit.html",
                context={"user_obj": user_obj, "error": "用户名已存在", "current_user": current_user},
                status_code=400
            )

        if new_password or confirm_password:
            if new_password != confirm_password:
                return templates.TemplateResponse(
                    request=request,
                    name="user_edit.html",
                    context={"user_obj": user_obj, "error": "两次输入的新密码不一致", "current_user": current_user},
                    status_code=400
                )
            user_obj.password_hash = hash_password(new_password)

        user_obj.username = username
        user_obj.is_active = (is_active == "true")
        user_obj.is_admin = (is_admin == "true")

        if current_user and current_user.id == user_obj.id:
            user_obj.is_active = True
            user_obj.is_admin = True
            request.session["user"] = user_obj.username

        db.commit()

        add_flash(request, f"用户 {user_obj.username} 已保存", "success")
        return RedirectResponse(url="/users", status_code=303)
    finally:
        db.close()


@app.post("/users/{user_id}/delete")
def user_delete(request: Request, user_id: int):
    db: Session = SessionLocal()
    try:
        redirect = require_admin(request, db)
        if redirect:
            return redirect

        current_user = get_current_user(request, db)
        user_obj = db.query(User).filter(User.id == user_id).first()

        if user_obj is not None and (current_user is None or user_obj.id != current_user.id):
            username = user_obj.username
            db.delete(user_obj)
            db.commit()
            add_flash(request, f"用户 {username} 已删除", "success")
        else:
            add_flash(request, "不能删除当前登录用户", "warning")

        return RedirectResponse(url="/users", status_code=303)
    finally:
        db.close()
