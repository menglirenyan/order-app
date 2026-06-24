import re
from datetime import datetime
from io import BytesIO
from pathlib import Path


def get_pillow_tools():
    try:
        from PIL import Image, ImageDraw, ImageFont, ImageOps
    except ImportError as exc:
        raise RuntimeError("服务器缺少 Pillow，无法处理图片。请先安装 requirements.txt 中的依赖。") from exc
    return Image, ImageDraw, ImageFont, ImageOps


def load_quote_font(size: int, bold: bool = False):
    _, _, ImageFont, _ = get_pillow_tools()
    candidates = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc" if bold else "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc" if bold else "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for path in candidates:
        if path and Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def parse_quote_number(value, default: float = 0.0) -> float:
    match = re.search(r"-?\d+(?:\.\d+)?", str(value or "").replace(",", "").strip())
    if not match:
        return default
    try:
        return float(match.group(0))
    except ValueError:
        return default


def format_quote_money(value: float) -> str:
    return f"{value:.2f}"


def normalize_quote_rows(rows):
    if not isinstance(rows, list):
        return []
    normalized = []
    for row in rows[:80]:
        if not isinstance(row, dict):
            continue
        name = str(row.get("name") or row.get("title") or "").strip()
        if not name:
            continue
        quantity = str(row.get("quantity") or "1").strip()[:20]
        unit_price = str(row.get("price") or row.get("unit_price") or "").strip()[:30]
        amount_value = parse_quote_number(quantity, 1.0) * parse_quote_number(unit_price)
        normalized.append({
            "name": name[:80],
            "image_url": str(row.get("image_url") or "").strip()[:500],
            "size": str(row.get("size") or "").strip()[:80],
            "quantity": quantity,
            "price": unit_price,
            "amount": format_quote_money(amount_value),
            "amount_value": amount_value,
        })
    return normalized


def resolve_static_image_path(image_url: str):
    if not image_url.startswith("/static/"):
        return None
    static_root = Path("app/static").resolve()
    candidate = (Path("app") / image_url.lstrip("/")).resolve()
    try:
        candidate.relative_to(static_root)
    except ValueError:
        return None
    return candidate if candidate.is_file() else None


def open_quote_image(image_url: str):
    local_path = resolve_static_image_path(image_url) if image_url else None
    if not local_path:
        return None
    Image, _, _, ImageOps = get_pillow_tools()
    try:
        with Image.open(BytesIO(local_path.read_bytes())) as source:
            source = ImageOps.exif_transpose(source).convert("RGB")
            return ImageOps.contain(source, (260, 180), method=Image.Resampling.LANCZOS)
    except Exception:
        return None


def text_width(draw, text, font):
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[2] - bbox[0]


def wrap_text(draw, text, font, max_width):
    if not text:
        return [""]
    lines = []
    current = ""
    for char in text:
        candidate = current + char
        if current and text_width(draw, candidate, font) > max_width:
            lines.append(current)
            current = char
        else:
            current = candidate
    if current:
        lines.append(current)
    return lines


def draw_wrapped_text(draw, text, xy, font, fill, max_width, line_height, max_lines=3):
    x, y = xy
    lines = wrap_text(draw, text, font, max_width)
    if len(lines) > max_lines:
        lines = lines[:max_lines]
        lines[-1] = lines[-1].rstrip("。,.， ") + "..."
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        y += line_height


def build_quote_image(rows):
    Image, ImageDraw, _, _ = get_pillow_tools()
    title_font = load_quote_font(32, bold=True)
    header_font = load_quote_font(17, bold=True)
    body_font = load_quote_font(16)
    small_font = load_quote_font(13)
    margin, table_width = 40, 1120
    col_widths = [150, 320, 160, 90, 150, 250]
    title_height, header_height, row_height, footer_height = 76, 48, 210, 70
    width = table_width + margin * 2
    height = margin + title_height + header_height + row_height * len(rows) + footer_height + margin
    image = Image.new("RGB", (width, height), "#f8fafc")
    draw = ImageDraw.Draw(image)
    draw.rectangle((margin, margin, margin + table_width, height - margin), fill="#ffffff", outline="#d1d5db")
    draw.text((margin + 20, margin + 18), "资料清单", font=title_font, fill="#111827")
    draw.text((margin + table_width - 230, margin + 31), datetime.now().strftime("%Y-%m-%d %H:%M"), font=small_font, fill="#6b7280")
    headers = ["名称", "图片", "尺寸", "数量", "参考数值", "参考合计"]
    y, x = margin + title_height, margin
    for index, header in enumerate(headers):
        draw.rectangle((x, y, x + col_widths[index], y + header_height), fill="#eef2ff", outline="#cbd5e1")
        draw.text((x + 12, y + 14), header, font=header_font, fill="#1f2937")
        x += col_widths[index]
    y += header_height
    for row in rows:
        x = margin
        for width_value in col_widths:
            draw.rectangle((x, y, x + width_value, y + row_height), fill="#ffffff", outline="#e5e7eb")
            x += width_value
        draw_wrapped_text(draw, row["name"], (margin + 12, y + 18), body_font, "#111827", col_widths[0] - 24, 24, 6)
        image_cell_x = margin + col_widths[0]
        thumb = open_quote_image(row["image_url"])
        if thumb:
            image.paste(thumb, (image_cell_x + (col_widths[1] - thumb.width) // 2, y + (row_height - thumb.height) // 2))
        else:
            bounds = (image_cell_x + 30, y + 24, image_cell_x + col_widths[1] - 30, y + row_height - 24)
            draw.rectangle(bounds, fill="#f3f4f6", outline="#d1d5db")
            draw.text((bounds[0] + 76, bounds[1] + 72), "无图", font=small_font, fill="#6b7280")
        size_x = margin + col_widths[0] + col_widths[1]
        draw_wrapped_text(draw, row["size"], (size_x + 12, y + 18), body_font, "#374151", col_widths[2] - 24, 24, 4)
        qty_x = size_x + col_widths[2]
        draw_wrapped_text(draw, row["quantity"], (qty_x + 12, y + 18), body_font, "#374151", col_widths[3] - 24, 24, 2)
        price_x = qty_x + col_widths[3]
        draw_wrapped_text(draw, row["price"], (price_x + 12, y + 18), body_font, "#374151", col_widths[4] - 24, 24, 3)
        amount_x = price_x + col_widths[4]
        draw_wrapped_text(draw, row["amount"], (amount_x + 12, y + 18), body_font, "#111827", col_widths[5] - 24, 24, 2)
        y += row_height
    total_value = sum(row.get("amount_value") or 0 for row in rows)
    draw.text((margin + 20, y + 18), "此资料清单由当前页面临时生成，系统内不保存。", font=small_font, fill="#6b7280")
    total_text = f"总计：{format_quote_money(total_value)}"
    draw.text((margin + table_width - text_width(draw, total_text, title_font) - 20, y + 14), total_text, font=title_font, fill="#dc2626")
    output = BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    return output


def build_quote_excel(rows):
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl，无法生成表格。请先安装 requirements.txt 中的依赖。") from exc
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "资料清单"
    sheet.append(["资料名称", "图片", "尺寸", "数量", "参考数值", "参考合计"])
    for column, width in {"A": 28, "B": 18, "C": 28, "D": 12, "E": 16, "F": 16}.items():
        sheet.column_dimensions[column].width = width
    header_fill = PatternFill("solid", fgColor="EEF2FF")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    image_refs = []
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, row in enumerate(rows, start=2):
        for column, value in enumerate([row["name"], "", row["size"], row["quantity"], row["price"], row["amount"]], start=1):
            sheet.cell(index, column, value)
        sheet.row_dimensions[index].height = 86
        for column in range(1, 7):
            cell = sheet.cell(index, column)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        thumb = open_quote_image(row["image_url"])
        if thumb:
            buffer = BytesIO()
            thumb.save(buffer, format="PNG")
            buffer.seek(0)
            excel_image = ExcelImage(buffer)
            excel_image.width = 88
            excel_image.height = 88
            sheet.add_image(excel_image, f"B{index}")
            image_refs.append(buffer)
    total_row = len(rows) + 2
    sheet.cell(total_row, 5, "总计")
    sheet.cell(total_row, 6, format_quote_money(sum(row.get("amount_value") or 0 for row in rows)))
    for column in range(1, 7):
        cell = sheet.cell(total_row, column)
        cell.border = border
        cell.font = Font(bold=True, color="111827")
    workbook._quote_image_refs = image_refs
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
